import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment

wb = Workbook()

FONT = "Arial"
NAVY = "1F3864"
LBLUE = "DCE6F1"
GREY = "F2F2F2"
WHITE = "FFFFFF"
BLUEFONT = "0000FF"
GREENFONT = "008000"
BLACKFONT = "000000"
YELLOWFILL = "FFFF00"

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name=FONT, size=size, bold=bold, color=color)
def title_font(size=16, bold=True, color=NAVY):
    return Font(name=FONT, size=size, bold=bold, color=color)
def lbl_font(size=10, bold=False, italic=False, color=BLACKFONT):
    return Font(name=FONT, size=size, bold=bold, italic=italic, color=color)
def input_font(size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold, color=BLUEFONT)
def formula_font(size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold, color=BLACKFONT)
def link_font(size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold, color=GREENFONT)

navy_fill = PatternFill("solid", fgColor=NAVY)
lblue_fill = PatternFill("solid", fgColor=LBLUE)
grey_fill = PatternFill("solid", fgColor=GREY)
yellow_fill = PatternFill("solid", fgColor=YELLOWFILL)

thin = Side(style="thin", color="B7B7B7")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
top_border = Border(top=Side(style="thin", color="000000"))
dbl_top = Border(top=Side(style="double", color="000000"))

CUR0 = '$#,##0;($#,##0);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'
NUM0 = '#,##0;(#,##0);"-"'
GWFMT = '#,##0.0"GW";(#,##0.0)"GW";"-"'

def sheet_header(ws, title, subtitle=None):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N1")
    c = ws["A1"]; c.value = title; c.font = title_font()
    if subtitle:
        ws.merge_cells("A2:N2")
        c2 = ws["A2"]; c2.value = subtitle; c2.font = lbl_font(italic=True, color="595959")

def section_row(ws, row, text, ncols=10, start_col=1):
    for i in range(ncols):
        ws.cell(row=row, column=start_col+i).fill = navy_fill
    cell = ws.cell(row=row, column=start_col)
    cell.value = text; cell.font = hdr_font()

def col_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i)
        c.value = h
        c.font = hdr_font(size=10, color="000000")
        c.fill = grey_fill if i == 0 else lblue_fill
        c.alignment = Alignment(horizontal="left" if i == 0 else "center")
        c.border = box

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def note(ws, cell_ref, text):
    ws[cell_ref].comment = Comment(text, "Model")

def L(i):
    """0-indexed -> column letter"""
    return get_column_letter(i+1)

# =========================================================================
# Period definitions
# =========================================================================
# Forecast quarters: Q3'26E .. Q4'28E  (10 quarters)
FQ = ["Q3'26E","Q4'26E","Q1'27E","Q2'27E","Q3'27E","Q4'27E","Q1'28E","Q2'28E","Q3'28E","Q4'28E"]
NF = len(FQ)
# Assumptions tab: col B is first forecast quarter -> columns B..(B+NF-1)
ASSUMP_COLS = [L(1+i) for i in range(NF)]   # B..K
# Segment Forecast tab: col B = Q2'26A actual, C.. = forecast quarters
SF_FCOLS = [L(2+i) for i in range(NF)]      # C..L
# Consolidated P&L / Dashboard: B,C,D = historical (Q2'25A, Q1'26A, Q2'26A); E.. = forecast
HIST3 = ["Q2'25A", "Q1'26A", "Q2'26A"]
ALLQ = HIST3 + FQ
CP_COLS = [L(i) for i in range(len(ALLQ))]  # B..N (13 cols total incl. header offset -> starts at col idx1=B)
# careful: L(0)='A' is label col, so data starts at L(1)='B'
CP_COLS = [L(1+i) for i in range(len(ALLQ))]  # B..N

print("FQ:", FQ)
print("ASSUMP_COLS:", ASSUMP_COLS)
print("SF_FCOLS:", SF_FCOLS)
print("CP_COLS:", CP_COLS)

# =========================================================================
# 1. COVER
# =========================================================================
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
set_col_widths(ws, [3, 46, 46, 3])
ws.merge_cells("B2:C2")
ws["B2"] = "Space Exploration Technologies Corp. (Nasdaq: SPCX)"
ws["B2"].font = Font(name=FONT, size=20, bold=True, color=NAVY)
ws.merge_cells("B3:C3")
ws["B3"] = "P&L Business Case Model"
ws["B3"].font = Font(name=FONT, size=14, bold=True, color="595959")
ws.merge_cells("B4:C4")
ws["B4"] = "Actuals through Q2 2026 with quarterly forecast build through Q4 2028"
ws["B4"].font = lbl_font(italic=True)

rows = [
    ("Prepared", "Claude (Anthropic) — created in claude.ai"),
    ("Currency / units", "US$ in millions unless noted"),
    ("Historical source", "SpaceX Q2 2026 earnings release (ir.spacex.com), with Q1 2026 and Q2 2025 comparatives disclosed therein"),
    ("", "SpaceX 'Reports Second Quarter 2026 Results', posted Aug 4, 2026 — https://s21.q4cdn.com/184289198/files/doc_financials/2026/q2/SpaceX-Reports-Second-Quarter-2026-Results.pdf"),
    ("Forecast basis", "Illustrative business-case projection. All forecast drivers are editable blue/yellow inputs on the Assumptions tab — this is NOT SpaceX guidance."),
    ("AI compute scenario", "AI segment forecast is built bottom-up off a nameplate compute (GW) build-out schedule: 1.4 GW at Q2'26A, 2.2 GW at Q4'26E (consistent with management's 'more than two gigawatts' comment), 10.0 GW at Q4'27E per business-case instruction, and an illustrative continued ramp to 15.0 GW by Q4'28E."),
    ("Switchable cases", "The Assumptions tab carries a case selector (1/2/3) governing AI capex per GW, blended revenue per GW and the marginal contract rate on new capacity. Case 1 is CoreWeave-anchored and sceptical; Case 2 reproduces management's $100B December-2026 ARR and ~1-year payback; Case 3 stresses Case 2 revenue against Nvidia's rising cost curve."),
    ("Company background", "SpaceX completed its IPO on June 12, 2026 (Nasdaq: SPCX), raising ~$85.7B net proceeds, and issued $25B of senior notes on June 26, 2026. Q2 2026 was its first quarter reporting as a public company."),
]
r = 6
for label, val in rows:
    ws.cell(row=r, column=2, value=label).font = lbl_font(bold=True)
    ws.cell(row=r, column=3, value=val).font = lbl_font()
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30 if val and len(val) > 90 else 16
    r += 2

r += 1
ws.cell(row=r, column=2, value="Contents").font = Font(name=FONT, size=12, bold=True, color=NAVY)
r += 1
contents = [
    "1. Assumptions — forecast drivers (editable inputs), including AI compute (GW) build-out",
    "2. KPIs — historical operating metrics (Starlink subs/ARPU, launches, AI compute)",
    "3. Historical P&L by Segment — Space / Connectivity / AI, Q2'25, Q1'26, Q2'26, H1'25, H1'26",
    "4. Segment Forecast Detail — quarterly build for each segment, Q3'26E-Q4'28E",
    "5. Consolidated P&L — historical actuals + quarterly forecast through Q4 2028",
    "6. Balance Sheet & Cash Flow — actuals at June 30, 2026 vs Dec 31, 2025",
    "7. Dashboard — key charts",
]
for line in contents:
    ws.cell(row=r, column=2, value=line).font = lbl_font(); r += 1

r += 1
ws.cell(row=r, column=2, value="Color legend:").font = lbl_font(bold=True); r += 1
ws.cell(row=r, column=2, value="Blue = hardcoded input / actual reported figure").font = input_font(); r += 1
ws.cell(row=r, column=2, value="Black = calculated formula").font = formula_font(); r += 1
ws.cell(row=r, column=2, value="Green = link from another sheet").font = link_font(); r += 1
c = ws.cell(row=r, column=2, value="Yellow fill = key assumption to edit"); c.fill = yellow_fill; c.font = lbl_font()

print("Cover done")
wb.save("/home/claude/spacex_model/spacex_model.xlsx")

# =========================================================================
# 2. ASSUMPTIONS
# =========================================================================
ws = wb.create_sheet("Assumptions")
sheet_header(ws, "Forecast Assumptions", "Editable inputs driving Segment Forecast and Consolidated P&L. AI segment is built off a nameplate compute (GW) schedule reaching 2.2 GW at Q4'26E, 10.0 GW at Q4'27E and 15.0 GW at Q4'28E. Set the case selector below before reading any output.")
set_col_widths(ws, [36] + [10]*NF + [3, 40])

row = 4
col_headers(ws, row, ["Driver"] + FQ)
row += 1

def input_row(ws, row, label, values, fmt=PCT1, fill=True):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=2+i, value=v)
        c.font = input_font(); c.number_format = fmt
        if fill: c.fill = yellow_fill
        c.border = box
    assert len(values) == NF, f"{label}: expected {NF} values, got {len(values)}"

section_row(ws, row, "CASE SELECTOR — type 1, 2 or 3 in the yellow cell", ncols=1+NF); row += 1
r_case = row
ws.cell(row=row, column=1, value="Active case (1 / 2 / 3)").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=1)
c.font = input_font(bold=True); c.number_format = NUM0; c.fill = yellow_fill; c.border = box
ws.cell(row=row, column=3, value=f'=CHOOSE($B${r_case},"CONSERVATIVE — CoreWeave-anchored","MANAGEMENT-CONSISTENT — reproduces $100B Dec-26 ARR","HIGH-CAPEX STRESS — management revenue, Nvidia cost curve")').font = formula_font(bold=True)
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=1+NF)
row += 1
for txt in [
 "Case 1 CONSERVATIVE — AI economics anchored to CoreWeave's disclosed contracted rates. Produces ~$68B ARR in Dec-2026 and a ~3 year capex payback: it does NOT reproduce management's guidance, and is the sceptical benchmark.",
 "Case 2 MANAGEMENT-CONSISTENT — blended revenue per GW raised toward the level implied by the $100B ARR guidance, capex per GW raised to the ~$50B cited by Nvidia and SemiAnalysis. Reproduces both the $100B Dec-2026 ARR and a ~1 year payback on new capital.",
 "Case 3 HIGH-CAPEX STRESS — case 2 revenue, but capex per GW follows Jensen Huang's projection of $50B today rising toward $90B. Tests whether the returns survive cost inflation.",
]:
    ws.cell(row=row, column=1, value=txt).font = lbl_font(size=8, italic=True, color="595959")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+NF)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    row += 1
row += 1

section_row(ws, row, "CASE LIBRARY — stored inputs for each case (the active rows below pick from these)", ncols=1+NF); row += 1
LIB = {}
lib_specs = [
 (1, 'blended', "Case 1: blended revenue per GW ($mm/qtr)", [2800,2900,2900,2850,2800,2750,2650,2550,2450,2400]),
 (1, 'marginal', "Case 1: marginal contract rate on NEW capacity ($mm/qtr)", [3200,3300,3300,3200,3150,3100,3000,2900,2800,2700]),
 (1, 'capex', "Case 1: capex per incremental GW ($mm/GW)", [32000,30000,28000,26500,25000,24000,23000,22000,21500,21000]),
 (2, 'blended', "Case 2: blended revenue per GW ($mm/qtr)", [6000,6500,6400,6300,6200,6100,5900,5700,5400,5200]),
 (2, 'marginal', "Case 2: marginal contract rate on NEW capacity ($mm/qtr)", [12000,12000,11800,11500,11200,11000,10500,10000,9500,9000]),
 (2, 'capex', "Case 2: capex per incremental GW ($mm/GW)", [48000,50000,52000,54000,56000,58000,60000,62000,64000,66000]),
 (3, 'blended', "Case 3: blended revenue per GW ($mm/qtr)", [6000,6500,6400,6300,6200,6100,5900,5700,5400,5200]),
 (3, 'marginal', "Case 3: marginal contract rate on NEW capacity ($mm/qtr)", [12000,12000,11800,11500,11200,11000,10500,10000,9500,9000]),
 (3, 'capex', "Case 3: capex per incremental GW ($mm/GW)", [55000,60000,65000,70000,75000,80000,84000,87000,89000,90000]),
]
for cs, metric, label, vals in lib_specs:
    LIB[(cs, metric)] = row
    input_row(ws, row, label, vals, fmt=CUR0, fill=False)
    row += 1
note(ws, f"B{LIB[(2,'marginal')]}", "SemiAnalysis puts the Google agreement at roughly $48B per gigawatt per year, and Altimeter's work shows xAI monetising at about 2-3x neocloud pricing. $12,000mm/quarter equals $48B/GW/yr. This is the rate NEW capacity earns, and is what management's sub-1-year payback claim refers to.")
note(ws, f"B{LIB[(2,'blended')]}", "The blended rate across the whole installed base is necessarily lower than the marginal rate, because older capacity carries earlier, cheaper contracts. $6,500mm/quarter at Q4'26 equals $26B/GW/yr, which is what reproduces the $100B December ARR guidance.")
note(ws, f"B{LIB[(3,'capex')]}", "Nvidia's Jensen Huang has projected capex per gigawatt rising from roughly $50B today toward $90B within a few years.")
row += 1

section_row(ws, row, "SPACE SEGMENT", ncols=1+NF); row += 1
r_space_growth = row; input_row(ws, row, "Revenue QoQ growth %",
    [0.10,0.12,0.10,0.09,0.08,0.08, 0.07,0.07,0.06,0.06]); row += 1
r_space_cogs = row; input_row(ws, row, "Cost of revenue, % of revenue",
    [0.33,0.32,0.31,0.30,0.30,0.29, 0.28,0.28,0.27,0.27]); row += 1
r_space_rd = row; input_row(ws, row, "R&D, % of revenue",
    [1.00,0.90,0.80,0.70,0.62,0.55, 0.50,0.46,0.42,0.39]); row += 1
r_space_sga = row; input_row(ws, row, "SG&A, % of revenue",
    [0.10,0.10,0.09,0.09,0.09,0.08, 0.08,0.08,0.07,0.07]); row += 1
r_space_capex = row; input_row(ws, row, "Capex, % of revenue",
    [1.00,0.90,0.75,0.65,0.55,0.50, 0.45,0.40,0.38,0.35]); row += 1

row += 1
section_row(ws, row, "CONNECTIVITY SEGMENT", ncols=1+NF); row += 1
r_conn_adds = row; input_row(ws, row, "Net Starlink subscriber adds (millions)",
    [1.6,1.6,1.5,1.5,1.4,1.4, 1.3,1.3,1.2,1.2], fmt=NUM1); row += 1
r_conn_arpu = row; input_row(ws, row, "Consumer ARPU ($/month)",
    [65,65,64,64,63,63, 62,62,61,61], fmt='$#,##0'); row += 1
r_conn_eg_growth = row; input_row(ws, row, "Enterprise & Government revenue QoQ growth %",
    [0.20,0.18,0.15,0.13,0.12,0.11, 0.10,0.10,0.09,0.09]); row += 1
r_conn_cogs = row; input_row(ws, row, "Cost of revenue, % of revenue",
    [0.47,0.46,0.45,0.44,0.43,0.42, 0.41,0.41,0.40,0.40]); row += 1
r_conn_rd = row; input_row(ws, row, "R&D, % of revenue",
    [0.065,0.062,0.060,0.058,0.056,0.055, 0.054,0.053,0.052,0.051]); row += 1
r_conn_sga = row; input_row(ws, row, "SG&A, % of revenue",
    [0.060,0.058,0.056,0.054,0.052,0.050, 0.049,0.048,0.047,0.046]); row += 1
r_conn_capex = row; input_row(ws, row, "Capex, % of revenue",
    [0.30,0.29,0.28,0.27,0.26,0.25, 0.24,0.24,0.23,0.23]); row += 1

row += 1
section_row(ws, row, "AI SEGMENT", ncols=1+NF); row += 1
r_ai_gw = row; input_row(ws, row, "Nameplate compute, period end (GW)",
    [1.8,2.2,3.0,4.5,7.0,10.0, 11.5,13.0,14.0,15.0], fmt=GWFMT); row += 1
r_ai_monetizable = row; input_row(ws, row, "Monetizable share of compute (%) — net of internal Grok training",
    [0.90,0.90,0.90,0.90,0.90,0.90, 0.90,0.90,0.90,0.90]); row += 1
r_ai_rev_per_gw = row
ws.cell(row=row, column=1, value="ACTIVE: blended revenue per GW ($mm/quarter)").font = lbl_font(bold=True)
for i, acol in enumerate(ASSUMP_COLS):
    cell = ws.cell(row=row, column=2+i,
        value=f"=CHOOSE($B${r_case},{acol}{LIB[(1,'blended')]},{acol}{LIB[(2,'blended')]},{acol}{LIB[(3,'blended')]})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = box
row += 1
r_ai_marginal_gw = row
ws.cell(row=row, column=1, value="ACTIVE: marginal rate on new capacity ($mm/quarter)").font = lbl_font(bold=True)
for i, acol in enumerate(ASSUMP_COLS):
    cell = ws.cell(row=row, column=2+i,
        value=f"=CHOOSE($B${r_case},{acol}{LIB[(1,'marginal')]},{acol}{LIB[(2,'marginal')]},{acol}{LIB[(3,'marginal')]})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = box
row += 1
r_grok_rev = row; input_row(ws, row, "Grok subscription revenue ($mm/qtr) [ASSUMPTION - unanchored]",
    [700,900,1150,1400,1650,1900, 2150,2400,2650,2900], fmt=CUR0); row += 1
r_cursor_rev = row; input_row(ws, row, "Cursor revenue ($mm/qtr) [ASSUMPTION - unanchored]",
    [0,1200,1500,1800,2100,2400, 2700,3000,3300,3600], fmt=CUR0); row += 1
r_ai_adv_growth = row; input_row(ws, row, "Advertising revenue QoQ growth %",
    [-0.02,0.00,0.01,0.02,0.02,0.02, 0.02,0.02,0.02,0.02]); row += 1
r_ai_cogs = row; input_row(ws, row, "Cost of revenue, % of revenue",
    [0.40,0.39,0.38,0.37,0.36,0.35, 0.34,0.34,0.33,0.33]); row += 1
r_ai_rd = row; input_row(ws, row, "R&D, % of revenue",
    [0.75,0.65,0.58,0.52,0.47,0.43, 0.40,0.37,0.35,0.33]); row += 1
r_ai_sga = row; input_row(ws, row, "SG&A, % of revenue",
    [0.18,0.16,0.15,0.14,0.13,0.12, 0.11,0.11,0.10,0.10]); row += 1
r_ai_capex_per_gw = row
ws.cell(row=row, column=1, value="ACTIVE: capex per incremental GW ($mm/GW)").font = lbl_font(bold=True)
for i, acol in enumerate(ASSUMP_COLS):
    cell = ws.cell(row=row, column=2+i,
        value=f"=CHOOSE($B${r_case},{acol}{LIB[(1,'capex')]},{acol}{LIB[(2,'capex')]},{acol}{LIB[(3,'capex')]})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = box
row += 1

row += 1
section_row(ws, row, "ORBITAL AI COMPUTE (SPECULATIVE — memo only, NOT included in Total Capex/Revenue/Net Income/Payback below)", ncols=1+NF); row += 1
r_orbital_gw = row; input_row(ws, row, "Orbital AI compute, period end (GW) — per SpaceX's own stated target",
    [0.0,0.0,0.0,0.1,0.4,1.0, 2.0,4.0,7.0,10.0], fmt=GWFMT); row += 1
r_orbital_capex_per_gw = row; input_row(ws, row, "Capex per incremental orbital GW ($mm/GW) — satellites + Starship launch, no data-center build",
    [55000,55000,54000,53000,52000,50000, 48000,46000,44000,42000], fmt=CUR0); row += 1
note(ws, f"B{r_orbital_gw}", "SpaceX targeted 1 GW of orbital AI compute by end of 2027, then 'order of magnitude' annual scaling thereafter, per its June 2026 investor disclosures (AI1 satellite unveiling, Gigasat factory). Musk: 'take this with a grain of salt.' First AI1 satellite launches not expected before early/mid-2027 — this schedule assumes a ramp starting then. Fully speculative; edit freely.")
note(ws, f"B{r_orbital_capex_per_gw}", "Anchored to the ~$55B total investment reported for the Gigasat/orbital program against its 1GW target (~$55B/GW) — higher than terrestrial capex/GW since it's an unproven, pre-scale approach (satellite manufacturing + launch cost, no comparable company data exists). Light decline assumes some mass-production/reusable-launch cost learning. There is NO real operating history for this line — treat it as illustrative only.")
row += 1

row += 1
section_short = "D&A AND FINANCING DRIVERS (feed the 'Financing & D&A' tab)"
section_row(ws, row, section_short, ncols=1+NF); row += 1

def single_input(ws, row, label, value, fmt, comment=None):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    c = ws.cell(row=row, column=2, value=value)
    c.font = input_font(); c.number_format = fmt; c.fill = yellow_fill; c.border = box
    if comment:
        note(ws, f"B{row}", comment)
    return row

r_it_share = single_input(ws, row, "IT equipment share of capex (%)", 0.70, PCT1,
    "Orennia and Epoch AI both put IT hardware (GPUs, networking, storage) at roughly 60-70% of AI data centre capex, with the balance in building/power/cooling."); row += 1
r_it_life = single_input(ws, row, "IT equipment depreciable life (quarters)", 20, NUM0,
    "20 quarters = 5 years, per Epoch AI's 1GW TCO model. Shortening this materially increases D&A."); row += 1
r_fac_life = single_input(ws, row, "Facility / infrastructure depreciable life (quarters)", 56, NUM0,
    "56 quarters = 14 years, per Epoch AI's 1GW TCO model."); row += 1
r_da_lag = single_input(ws, row, "Quarters from capex spend to depreciation start", 2, NUM0,
    "Matches the 6-month build/ramp lag used in the AI capex payback check — assets are not depreciated until placed in service."); row += 1
r_legacy_noncash = single_input(ws, row, "Legacy non-cash addback ($mm/quarter)", 3681, CUR0,
    "Q2'26A total company Adjusted EBITDA ($3,538mm) less operating loss (-$143mm) = $3,681mm of D&A, SBC and other non-cash charges already embedded in the reported actuals — and therefore already embedded in the cost ratios used in this forecast. Held flat so that only NEW-capex depreciation is added on top (avoids double-counting)."); row += 1
r_exist_debt = single_input(ws, row, "Existing debt at Jun 30, 2026 ($mm)", 39364, CUR0,
    "Balance sheet: $2,525mm current + $36,839mm non-current debt and finance leases."); row += 1
r_exist_rate = single_input(ws, row, "Interest rate on existing debt (%/yr)", 0.055, PCT1,
    "Blended estimate. The $25B senior notes issued Jun 26, 2026 carried a 5.855% weighted average rate; older facilities assumed slightly lower."); row += 1
r_new_rate = single_input(ws, row, "Interest rate on NEW debt raised (%/yr)", 0.065, PCT1,
    "Assumed premium to existing cost of debt given the scale of incremental issuance required."); row += 1
r_cash_yield = single_input(ws, row, "Yield on cash & marketable securities (%/yr)", 0.040, PCT1,
    "Drives interest income, replacing the previous flat per-quarter assumption."); row += 1
r_min_cash = single_input(ws, row, "Minimum cash buffer ($mm)", 15000, CUR0,
    "New debt is drawn only to the extent cash would otherwise fall below this floor."); row += 1
r_dec_share = single_input(ws, row, "Exit-month share of quarterly revenue (%)", 0.38, PCT1,
    "Management defines ARR as expected December revenue annualised, not the quarter averaged. Revenue ramps within the quarter — the $6.7B of cloud services contracted in early Q3 'begins ramping starting in October' — so the final month runs above a flat 33.3%. At 38% the exit month annualises to 4.56x quarterly revenue rather than 4.0x."); row += 1
r_opening_cash = single_input(ws, row, "Opening cash & securities at Jun 30, 2026 ($mm)", 100009, CUR0,
    "Balance sheet: $93,522mm cash and equivalents + $6,487mm marketable securities."); row += 1

row += 1
section_row(ws, row, "BELOW-THE-LINE / CONSOLIDATED", ncols=1+NF); row += 1
r_int_exp = row; input_row(ws, row, "Interest expense (per quarter, $mm)",
    [-680,-700,-700,-690,-680,-670, -660,-650,-640,-630], fmt=CUR0); row += 1
r_int_inc = row; input_row(ws, row, "Interest income (per quarter, $mm)",
    [700,680,660,640,630,620, 610,600,590,580], fmt=CUR0); row += 1
r_other_inc = row; input_row(ws, row, "Other income (expense), net (per quarter, $mm)",
    [-50,-50,-30,-20,-10,0, 0,0,0,0], fmt=CUR0); row += 1
r_tax = row; input_row(ws, row, "Effective tax adjustment (per quarter, $mm)",
    [20,20,20,20,20,20, 20,20,20,20], fmt=CUR0); row += 1

row += 1
section_row(ws, row, "MARKET CAP COMPARABLES (CLOUD COMPUTING / AI INFRASTRUCTURE)", ncols=1+NF); row += 1
col_headers(ws, row, ["Comparable company", "Market cap ($B)", "LTM revenue ($B)", "Mkt Cap / Revenue"], start_col=1)
row += 1
comps_start = row
comps = [
    ("CoreWeave (CRWV)", 44.0, 9.4, "Multiples.vc, market cap & LTM revenue as of Jul 24, 2026"),
    ("Nebius Group (NBIS)", 48.3, 0.878, "Yahoo Finance, market cap & TTM revenue as of Jul 31, 2026"),
    ("Oracle (ORCL)", 408.6, 67.35, "CompaniesMarketCap.com, market cap & TTM revenue as of Aug 2026"),
]
for name, mcap, rev, src in comps:
    ws.cell(row=row, column=1, value=name).font = lbl_font()
    c1 = ws.cell(row=row, column=2, value=mcap); c1.font = input_font(); c1.number_format = '$#,##0.0"B"'; c1.border = box
    c2 = ws.cell(row=row, column=3, value=rev); c2.font = input_font(); c2.number_format = '$#,##0.0"B"'; c2.border = box
    c3 = ws.cell(row=row, column=4, value=f"=B{row}/C{row}"); c3.font = formula_font(); c3.number_format = '0.0"x"'; c3.border = box
    note(ws, f"A{row}", src)
    row += 1
comps_end = row - 1
r_median_mult = row
ws.cell(row=row, column=1, value="Median Mkt Cap / Revenue multiple").font = lbl_font(bold=True)
c = ws.cell(row=row, column=4, value=f"=MEDIAN(D{comps_start}:D{comps_end})"); c.font = formula_font(bold=True); c.number_format = '0.0"x"'; c.border = top_border
row += 1
ws.cell(row=row, column=1, value="Average Mkt Cap / Revenue multiple").font = lbl_font()
c = ws.cell(row=row, column=4, value=f"=AVERAGE(D{comps_start}:D{comps_end})"); c.font = formula_font(); c.number_format = '0.0"x"'
row += 2
r_selected_mult = row
ws.cell(row=row, column=1, value="Selected Mkt Cap / Revenue multiple (applied to SpaceX LTM revenue)").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=f"=D{r_median_mult}"); c.font = input_font(bold=True); c.number_format = '0.0"x"'; c.fill = yellow_fill; c.border = box
note(ws, f"B{row}", "Defaults to the median comp multiple; overwrite with any number to test a different valuation scenario. NBIS trades at an extreme premium (LTM revenue is still small relative to its contracted backlog), so median (not average) is used as the base case to avoid skew.")
row += 2

section_row(ws, row, "PEG RATIO COMPARABLES (AI / MEGACAP TECH)", ncols=1+NF); row += 1
col_headers(ws, row, ["Comparable company", "PEG ratio (5yr expected)", "", ""], start_col=1)
row += 1
peg_comps_start = row
peg_comps = [
    ("NVIDIA (NVDA)", 0.55, "Yahoo Finance, PEG Ratio (5yr expected), as of Jul 31, 2026"),
    ("Palantir (PLTR)", 1.76, "Yahoo Finance, PEG Ratio (5yr expected), as of Jul 31, 2026"),
    ("Microsoft (MSFT)", 1.52, "Yahoo Finance, PEG Ratio (5yr expected), as of Jul 31, 2026"),
]
for name, peg, src in peg_comps:
    ws.cell(row=row, column=1, value=name).font = lbl_font()
    c1 = ws.cell(row=row, column=2, value=peg); c1.font = input_font(); c1.number_format = '0.00"x"'; c1.border = box
    note(ws, f"A{row}", src)
    row += 1
peg_comps_end = row - 1
r_median_peg = row
ws.cell(row=row, column=1, value="Median comparable PEG ratio").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=f"=MEDIAN(B{peg_comps_start}:B{peg_comps_end})"); c.font = formula_font(bold=True); c.number_format = '0.00"x"'; c.border = top_border
row += 1
ws.cell(row=row, column=1, value="Average comparable PEG ratio").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"=AVERAGE(B{peg_comps_start}:B{peg_comps_end})"); c.font = formula_font(); c.number_format = '0.00"x"'
row += 2
r_selected_peg = row
ws.cell(row=row, column=1, value="Selected PEG ratio (applied to SpaceX's own LTM net income growth)").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=f"=B{r_median_peg}"); c.font = input_font(bold=True); c.number_format = '0.00"x"'; c.fill = yellow_fill; c.border = box
note(ws, f"B{row}", "Defaults to the median of NVDA/PLTR/MSFT PEG ratios; overwrite to test a different peer benchmark. CoreWeave/Nebius are excluded here since their earnings are not yet stable enough for a meaningful PEG.")
row += 2

section_row(ws, row, "AI ASSUMPTIONS — EXTERNAL BENCHMARK CHECK (memo, not used in calculations)", ncols=1+NF); row += 1
bench_rows = [
    ("AI capex per GW ($B, Q3'26E -> Q4'28E)",
     "This model (recalibrated): $32.0B -> $21.0B/GW.  External comps: $20-60B/GW — Bernstein $35B, Nvidia $50-60B, Epoch AI $38B, Orennia $60B, Nebius FY26 guidance implies ~$22.5B, CoreWeave FY26 guidance implies ~$46B. Recalibrated to sit near the Bernstein/Epoch AI band and meaningfully closer to CoreWeave — deliberately below CoreWeave's own ~$46B/GW per the user's instruction.",
     "RECALIBRATED"),
    ("AI revenue per GW ($B/yr, Q3'26E -> Q4'28E)",
     "This model (recalibrated): $8.0B -> $5.2B/GW/yr.  External comps: SpaceX's own Q2'26A actual is ~$6.3B/GW/yr; CoreWeave's FY26 exit-ARR guidance ($18-19B) / exit active power (1.7GW) implies ~$10.9B/GW/yr. Recalibrated above SpaceX's own actual and roughly 2/3 of the way to CoreWeave's rate — deliberately below CoreWeave per the user's instruction.",
     "RECALIBRATED"),
    ("AI cost of revenue, % of revenue (Q3'26E -> Q4'28E)",
     "This model: 40% -> 33% (unchanged).  External comps: CoreWeave Q1'26 gross margin 65-69%, i.e. implied COGS ~31-35% of revenue.",
     "IN LINE"),
    ("AI R&D, % of revenue (Q3'26E -> Q4'28E)",
     "This model: 75% -> 33% (unchanged).  External comps: SpaceX's own Q2'26A actual AI R&D ratio was 2,178/2,561 = ~85% of revenue.",
     "IN LINE (if anything conservative)"),
]
for metric, detail, flag in bench_rows:
    flag_color = "1F6F1F" if flag == "RECALIBRATED" else ("C00000" if flag.startswith("OUT OF LINE") or flag.startswith("LIKELY UNDERSTATED") else "008000")
    c0 = ws.cell(row=row, column=1, value=f"{metric}  —  {flag}")
    c0.font = lbl_font(bold=True, color=flag_color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+NF)
    row += 1
    c1 = ws.cell(row=row, column=1, value=detail)
    c1.font = lbl_font(size=9)
    c1.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+NF)
    ws.row_dimensions[row].height = 28
    row += 1
row += 1
ws.cell(row=row, column=1, value="Sources: Bernstein/Investing.com (Nov 2025); Epoch AI 1GW TCO model (May 2026); Orennia 'What It Costs to Build a 1GW Data Center' (May 2026); Nebius FY26 capex guidance raised to $20-25B (Aug 2026); CoreWeave FY26 guidance of $30-35B capex / 1.7GW active power and $18-19B exit ARR (May-Aug 2026); CoreWeave Q1'26 10-Q gross margin. Capex/GW and revenue/GW were both recalibrated upward from the original assumptions on Aug 2026 at user request, to sit meaningfully closer to CoreWeave's real-world economics without matching them exactly. See the Segment Forecast tab's marginal payback check for the resulting effect on implied AI capex payback.").font = lbl_font(size=8, italic=True, color="595959")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+NF)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

row += 2
ws.cell(row=row, column=1, value="Notes").font = lbl_font(bold=True); row += 1
notes_txt = [
 "- Q2'26 actual results are the jump-off point. AI segment revenue = AI infrastructure revenue (revenue-generating GW x blended $/GW/quarter) + Grok subscription revenue + Cursor revenue + Advertising revenue.",
 "- Compute build-out: 1.4 GW at Q2'26A -> 2.2 GW at Q4'26E (consistent with management's 'more than two gigawatts' end-2026 comment) -> 10.0 GW at Q4'27E (per business-case instruction) -> illustrative continued ramp to 15.0 GW by Q4'28E. Edit the GW row directly to test other build-out scenarios.",
 "- Monetizable share of compute is set to 90%, reflecting management's statement that roughly 10% of compute is used for internal Grok training and therefore earns no external revenue. Capex is incurred on ALL GW built while only 90% earns revenue, so this both reduces AI revenue and lengthens the implied capex payback. The Q2'26A anchor column is left at 100% because the reported actual already nets out any internal usage in that quarter.",
 "- AI capex = incremental GW added in the quarter x $/GW build cost (compute hardware + data center + cooling + power infrastructure). RECALIBRATED against external comps (see benchmark table below): starts at $32B/GW in Q3'26E, above the Q1-Q2'26A actual blended run-rate (~$18-24B/GW) and closer to (but below) CoreWeave's implied ~$46B/GW and Nvidia's own $50-60B/GW figure, tapering only gently to $21B/GW by Q4'28E — external sources describe per-GW build costs as flat-to-rising, not sharply falling, so the taper here is much shallower than the original assumption.",
 "- Space and Connectivity capex are modeled as a % of segment revenue, benchmarked to the Q2'26A actual ratios (Space ~122% of revenue given heavy Starship investment; Connectivity ~32%) and tapering over the forecast.",
 "- AI infrastructure revenue per GW: RECALIBRATED against external comps (see benchmark table below): starts at $8.0B/GW/yr in Q3'26E — above SpaceX's own Q2'26A actual (~$6.3B/GW/yr) and closer to (but below) CoreWeave's implied ~$10.9B/GW/yr exit run-rate — declining to $5.2B/GW/yr by Q4'28E as installed capacity growth outpaces near-term contracted utilization. Still a simplifying assumption, not a contractual schedule.",
 "- Space: growth assumes accelerating Starship V3 cadence into 2027-28; R&D % of revenue eases as the vehicle matures.",
 "- Connectivity: modeled bottom-up from subscriber adds x ARPU (Consumer) plus a separate Enterprise & Government growth rate.",
 "- Below-the-line items are simplified straight-line assumptions given SpaceX does not guide interest/other income/tax; they are NOT derived from a debt/cash schedule. Edit freely.",
 "- GUIDANCE CHECK: SpaceX CFO Bret Johnsen stated on the Q2'26 earnings call that AI compute capex is generating 'less than a one-year payback.' The Segment Forecast tab computes an implied MARGINAL payback for every quarter (actual and forecast): capex committed in quarter X is assumed to come online and start earning revenue only in quarter X+2 (a 6-month build/ramp lag, with zero revenue during that window), monetized at the vintage-specific $/GW rate the model assumes will prevail in quarter X+2 — not a blended, whole-portfolio average. Under this basis, implied payback runs ~4.0-8.7 years (worse than a naive same-quarter blended calc, since the model's $/GW revenue rate declines over time and the lag pushes recovery further out) — the default assumptions do NOT reproduce management's <1-year guidance under either method. Either the $/GW revenue rate is set too low, the $/GW capex rate is set too high, or (most likely) management's payback is computed off specific contracted cash economics for new deals (e.g. the Google/Anthropic/Reflection AI agreements) rather than the company-wide average rate this model uses. Raise 'AI infrastructure revenue per GW' and/or lower 'Capex per incremental GW' to test what assumptions would be needed to hit <1 year.",
 "- ORBITAL AI COMPUTE (new section above): tracks SpaceX's separately-disclosed orbital data center ambition (AI1 satellites, Gigasat factory) — 1GW by end of 2027 per SpaceX's own June 2026 target, scaling 'order of magnitude' annually thereafter. This is ADDITIVE to (not a substitute for) the terrestrial GW build-out above, uses a completely different cost structure (satellites + Starship launches, not data centers), and has NO revenue or capex history to calibrate against — it does not flow into Total Capex, Total AI Revenue, Consolidated P&L, Net Income, Market Cap, or the AI capex payback check elsewhere in this model. It exists purely as a memo so the model shows SpaceX's full stated ambition alongside what's actually been reported and spent.",
 "- This is an illustrative business-case model, not SpaceX guidance or an analyst consensus estimate.",
]
for t in notes_txt:
    ws.cell(row=row, column=1, value=t).font = lbl_font(size=9, italic=True, color="595959")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+NF)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    row += 1

ws.freeze_panes = "B5"
ASSUMP_ROWS = dict(space_growth=r_space_growth, space_cogs=r_space_cogs, space_rd=r_space_rd, space_sga=r_space_sga, space_capex=r_space_capex,
    conn_adds=r_conn_adds, conn_arpu=r_conn_arpu, conn_eg_growth=r_conn_eg_growth, conn_cogs=r_conn_cogs, conn_rd=r_conn_rd, conn_sga=r_conn_sga, conn_capex=r_conn_capex,
    ai_gw=r_ai_gw, ai_monetizable=r_ai_monetizable, ai_rev_per_gw=r_ai_rev_per_gw,
    grok_rev=r_grok_rev, cursor_rev=r_cursor_rev, dec_share=r_dec_share,
    ai_marginal_gw=r_ai_marginal_gw, case=r_case, ai_adv_growth=r_ai_adv_growth, ai_cogs=r_ai_cogs, ai_rd=r_ai_rd, ai_sga=r_ai_sga, ai_capex_per_gw=r_ai_capex_per_gw,
    orbital_gw=r_orbital_gw, orbital_capex_per_gw=r_orbital_capex_per_gw,
    it_share=r_it_share, it_life=r_it_life, fac_life=r_fac_life, da_lag=r_da_lag,
    legacy_noncash=r_legacy_noncash, exist_debt=r_exist_debt, exist_rate=r_exist_rate,
    new_rate=r_new_rate, cash_yield=r_cash_yield, min_cash=r_min_cash, opening_cash=r_opening_cash,
    int_exp=r_int_exp, int_inc=r_int_inc, other_inc=r_other_inc, tax=r_tax,
    median_mult=r_median_mult, selected_mult=r_selected_mult, median_peg=r_median_peg, selected_peg=r_selected_peg)
print("Assumptions done", ASSUMP_ROWS)
wb.save("/home/claude/spacex_model/spacex_model.xlsx")

# =========================================================================
# 3. KPI SHEET (historical)
# =========================================================================
ws = wb.create_sheet("KPIs")
sheet_header(ws, "Operating KPIs (Actuals)", "Source: SpaceX Q2 2026 earnings release, segment operating & financial data tables. Forecast compute build-out is on the Assumptions tab.")
set_col_widths(ws, [34, 12, 12, 12, 3, 40])

row = 4
col_headers(ws, row, ["Metric", "Q2'25A", "Q1'26A", "Q2'26A"])
row += 1
section_row(ws, row, "CONNECTIVITY / STARLINK", ncols=4); row += 1
for label, a, b, c, fmt in [
    ("Starlink subscribers (millions, period end)", 6.0, 10.3, 12.0, NUM1),
    ("Starlink ARPU ($/month)", 85, 66, 66, '$#,##0'),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b, c]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = fmt; cell.border = box
    row += 1

row += 1
section_row(ws, row, "SPACE / LAUNCH", ncols=4); row += 1
for label, a, b, c, fmt in [
    ("Total launches (#)", 46, 40, 38, NUM0),
    ("Customer launches (#)", 9, 7, 10, NUM0),
    ("Internal launches (#)", 37, 33, 28, NUM0),
    ("Customer payloads (metric tons)", 88, 45, 87, NUM0),
    ("Internal payloads (metric tons)", 563, 511, 397, NUM0),
    ("Mass to orbit (metric tons)", 652, 556, 485, NUM0),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b, c]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = fmt; cell.border = box
    row += 1

row += 1
section_row(ws, row, "AI", ncols=4); row += 1
ws.cell(row=row, column=1, value="Nameplate compute (GW)").font = lbl_font()
for i, v in enumerate([0.4, 1.0, 1.4]):
    cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = '0.0'; cell.border = box
row += 1

row += 2
ws.cell(row=row, column=1, value="Forward compute build-out (Q3'26E-Q4'28E) is modeled on the Assumptions tab and feeds the Segment Forecast tab.").font = lbl_font(size=9, italic=True, color="595959")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("KPI sheet done")

# =========================================================================
# 4. HISTORICAL P&L BY SEGMENT
# =========================================================================
ws = wb.create_sheet("Historical P&L")
sheet_header(ws, "Historical P&L by Segment (Actuals)", "$ in millions. Source: SpaceX Q2 2026 earnings release (three- and six-month tables)")
set_col_widths(ws, [34, 12, 12, 12, 12, 12, 3])

HCOLS = ["B", "C", "D", "E", "F"]  # Q2'25A, Q1'26A, Q2'26A, H1'25A, H1'26A
row = 4
col_headers(ws, row, ["$mm", "Q2'25A", "Q1'26A", "Q2'26A", "H1'25A", "H1'26A"])
row += 1

def hist_block(ws, row, title, data_rows):
    section_row(ws, row, title, ncols=6)
    row += 1
    locs = {}
    for label, vals, fmt, kind in data_rows:
        c0 = ws.cell(row=row, column=1, value=label)
        is_subtotal = kind in ("sum2", "op")
        c0.font = lbl_font(bold=is_subtotal)
        if kind == "input":
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=2+i, value=v)
                cell.number_format = fmt; cell.font = input_font(); cell.border = box
        elif kind == "sum2":
            for i, col in enumerate(HCOLS):
                cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{row-2}:{col}{row-1})")
                cell.number_format = fmt; cell.font = formula_font(bold=True); cell.border = top_border
        elif kind == "op":
            for i, col in enumerate(HCOLS):
                cell = ws.cell(row=row, column=2+i, value=0)
                cell.number_format = fmt; cell.font = formula_font(bold=True); cell.border = top_border
        locs[label] = row
        row += 1
    return row, locs

row, space_locs = hist_block(ws, row, "SPACE SEGMENT", [
    ("Launch services revenue", [490, 330, 648, 1056, 978], CUR0, "input"),
    ("Launch & development revenue", [256, 289, 314, 555, 603], CUR0, "input"),
    ("Space revenue", None, CUR0, "sum2"),
    ("Cost of revenue", [330, 281, 329, 627, 610], CUR0, "input"),
    ("Research & development", [693, 930, 1076, 1219, 2006], CUR0, "input"),
    ("Selling, general & administrative", [87, 70, 99, 175, 169], CUR0, "input"),
    ("Impairment", [5, 0, 0, 29, 0], CUR0, "input"),
    ("Income (loss) from operations", None, CUR0, "op"),
    ("Adjusted EBITDA", [-93, -351, -205, 131, -556], CUR0, "input"),
    ("Capex", [946, 1052, 1174, 1705, 2226], CUR0, "input"),
])
space_rev_row = space_locs["Space revenue"]; space_op_row = space_locs["Income (loss) from operations"]
cost_rows = [space_locs["Cost of revenue"], space_locs["Research & development"], space_locs["Selling, general & administrative"], space_locs["Impairment"]]
for i, col in enumerate(HCOLS):
    cost_str = "".join([f"-{col}{cr}" for cr in cost_rows])
    ws.cell(row=space_op_row, column=2+i, value=f"={col}{space_rev_row}{cost_str}")
row += 1

row, conn_locs = hist_block(ws, row, "CONNECTIVITY SEGMENT", [
    ("Consumer revenue", [1721, 2148, 2485, 3213, 4633], CUR0, "input"),
    ("Enterprise & government revenue", [867, 1109, 1806, 1849, 2915], CUR0, "input"),
    ("Connectivity revenue", None, CUR0, "sum2"),
    ("Cost of revenue", [1401, 1651, 2060, 2615, 3711], CUR0, "input"),
    ("Research & development", [143, 205, 294, 266, 499], CUR0, "input"),
    ("Selling, general & administrative", [121, 213, 281, 225, 494], CUR0, "input"),
    ("Income from operations", None, CUR0, "op"),
    ("Adjusted EBITDA", [1583, 2087, 2597, 3200, 4684], CUR0, "input"),
    ("Capex", [1130, 1332, 1367, 1944, 2699], CUR0, "input"),
])
conn_rev_row = conn_locs["Connectivity revenue"]; conn_op_row = conn_locs["Income from operations"]
cost_rows = [conn_locs["Cost of revenue"], conn_locs["Research & development"], conn_locs["Selling, general & administrative"]]
for i, col in enumerate(HCOLS):
    cost_str = "".join([f"-{col}{cr}" for cr in cost_rows])
    ws.cell(row=conn_op_row, column=2+i, value=f"={col}{conn_rev_row}{cost_str}")
row += 1

row, ai_locs = hist_block(ws, row, "AI SEGMENT", [
    ("Advertising revenue", [426, 343, 367, 870, 710], CUR0, "input"),
    ("AI solutions & infrastructure revenue", [311, 475, 2194, 595, 2669], CUR0, "input"),
    ("AI revenue", None, CUR0, "sum2"),
    ("Cost of revenue", [551, 456, 1106, 1002, 1562], CUR0, "input"),
    ("Research & development", [1122, 2379, 2178, 2030, 4557], CUR0, "input"),
    ("Selling, general & administrative", [398, 463, 532, 699, 995], CUR0, "input"),
    ("Restructuring charges (credits)", [190, -11, 2, 194, -9], CUR0, "input"),
    ("Loss from operations", None, CUR0, "op"),
    ("Adjusted EBITDA", [-276, -609, 1146, -387, 537], CUR0, "input"),
    ("Capex", [749, 7723, 15828, 3316, 23551], CUR0, "input"),
])
ai_rev_row = ai_locs["AI revenue"]; ai_op_row = ai_locs["Loss from operations"]
cost_rows = [ai_locs["Cost of revenue"], ai_locs["Research & development"], ai_locs["Selling, general & administrative"], ai_locs["Restructuring charges (credits)"]]
for i, col in enumerate(HCOLS):
    cost_str = "".join([f"-{col}{cr}" for cr in cost_rows])
    ws.cell(row=ai_op_row, column=2+i, value=f"={col}{ai_rev_row}{cost_str}")
row += 1

section_row(ws, row, "TOTAL COMPANY (CHECK)", ncols=6); row += 1
tot_rev_row = row
ws.cell(row=row, column=1, value="Total revenue").font = formula_font(bold=True)
for i, col in enumerate(HCOLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{space_rev_row}+{col}{conn_rev_row}+{col}{ai_rev_row}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
tot_op_row = row
ws.cell(row=row, column=1, value="Total income (loss) from operations").font = formula_font(bold=True)
for i, col in enumerate(HCOLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{space_op_row}+{col}{conn_op_row}+{col}{ai_op_row}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1

note(ws, f"B{tot_rev_row}", "Should tie to reported consolidated revenue of $4,071mm (Q2'25), $4,694mm (Q1'26), $7,814mm (Q2'26)")
HIST_ROWS = dict(space_rev=space_rev_row, space_op=space_op_row, conn_rev=conn_rev_row, conn_op=conn_op_row,
                  ai_rev=ai_rev_row, ai_op=ai_op_row, tot_rev=tot_rev_row, tot_op=tot_op_row,
                  space_locs=space_locs, conn_locs=conn_locs, ai_locs=ai_locs)
ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Historical P&L sheet done", HIST_ROWS)

# =========================================================================
# 5. SEGMENT FORECAST DETAIL
# =========================================================================
ws = wb.create_sheet("Segment Forecast")
sheet_header(ws, "Segment Forecast Detail (Illustrative Business Case)", "$ in millions unless noted. Q2'26A = actual jump-off quarter. Q3'26E-Q4'28E driven by Assumptions tab. AI revenue = (revenue-generating GW x blended $/GW/quarter) + Grok subscriptions + Cursor + Advertising.")
set_col_widths(ws, [36] + [11]*(1+NF) + [3])

row = 4
col_headers(ws, row, ["$mm"] + ["Q2'26A"] + FQ)
row += 1

def asum_ref(r, acol):
    return f"Assumptions!{acol}{r}"
def hist_ref(cellrow):
    return f"'Historical P&L'!D{cellrow}"  # col D = Q2'26A on Historical P&L

def cost_line(ws, row, label, assum_row, rev_row_local):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, col in enumerate(SF_FCOLS):
        acol = ASSUMP_COLS[i]
        cell = ws.cell(row=row, column=3+i, value=f"={col}{rev_row_local}*{asum_ref(assum_row, acol)}")
        cell.font = formula_font(); cell.number_format = CUR0
    return row

# --- SPACE ---
section_row(ws, row, "SPACE SEGMENT", ncols=2+NF); row += 1
sf_space_rev = row
ws.cell(row=row, column=1, value="Revenue").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(HIST_ROWS['space_rev'])}"); c.font = link_font(bold=True); c.number_format = CUR0
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={prev_col}{row}*(1+{asum_ref(ASSUMP_ROWS['space_growth'], acol)})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0
    prev_col = col
row += 1
ws.cell(row=row, column=1, value="  QoQ growth %").font = lbl_font(italic=True, size=9)
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['space_growth'], acol)}")
    cell.font = link_font(size=9); cell.number_format = PCT1
row += 1
sf_space_cogs = row = cost_line(ws, row, "Cost of revenue", ASSUMP_ROWS['space_cogs'], sf_space_rev); row += 1
sf_space_rd = row = cost_line(ws, row, "Research & development", ASSUMP_ROWS['space_rd'], sf_space_rev); row += 1
sf_space_sga = row = cost_line(ws, row, "Selling, general & administrative", ASSUMP_ROWS['space_sga'], sf_space_rev); row += 1
sf_space_op = row
ws.cell(row=row, column=1, value="Income (loss) from operations").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(HIST_ROWS['space_op'])}"); c.font = link_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_space_rev}-{col}{sf_space_cogs}-{col}{sf_space_rd}-{col}{sf_space_sga}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
sf_space_capex = row = cost_line(ws, row, "Capex", ASSUMP_ROWS['space_capex'], sf_space_rev)
c = ws.cell(row=sf_space_capex, column=2, value=f"={hist_ref(space_locs['Capex'])}"); c.font = link_font(); c.number_format = CUR0
row += 2

# --- CONNECTIVITY ---
section_row(ws, row, "CONNECTIVITY SEGMENT", ncols=2+NF); row += 1
sf_subs = row
ws.cell(row=row, column=1, value="Starlink subscribers, period end (mm)").font = lbl_font()
c = ws.cell(row=row, column=2, value=12.0); c.font = input_font(); c.number_format = NUM1
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={prev_col}{row}+{asum_ref(ASSUMP_ROWS['conn_adds'], acol)}")
    cell.font = formula_font(); cell.number_format = NUM1
    prev_col = col
row += 1
sf_arpu = row
ws.cell(row=row, column=1, value="Consumer ARPU ($/month)").font = lbl_font()
c = ws.cell(row=row, column=2, value=66); c.font = input_font(); c.number_format = '$#,##0'
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['conn_arpu'], acol)}")
    cell.font = link_font(); cell.number_format = '$#,##0'
row += 1
sf_consumer_rev = row
ws.cell(row=row, column=1, value="Consumer revenue (subs x ARPU x 3 months)").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"={hist_ref(conn_locs['Consumer revenue'])}"); c.font = link_font(); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_subs}*{col}{sf_arpu}*3")
    cell.font = formula_font(); cell.number_format = CUR0
row += 1
sf_eg_rev = row
ws.cell(row=row, column=1, value="Enterprise & government revenue").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"={hist_ref(conn_locs['Enterprise & government revenue'])}"); c.font = link_font(); c.number_format = CUR0
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={prev_col}{row}*(1+{asum_ref(ASSUMP_ROWS['conn_eg_growth'], acol)})")
    cell.font = formula_font(); cell.number_format = CUR0
    prev_col = col
row += 1
sf_conn_rev = row
ws.cell(row=row, column=1, value="Connectivity revenue").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(HIST_ROWS['conn_rev'])}"); c.font = link_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_consumer_rev}+{col}{sf_eg_rev}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
sf_conn_cogs = row = cost_line(ws, row, "Cost of revenue", ASSUMP_ROWS['conn_cogs'], sf_conn_rev); row += 1
sf_conn_rd = row = cost_line(ws, row, "Research & development", ASSUMP_ROWS['conn_rd'], sf_conn_rev); row += 1
sf_conn_sga = row = cost_line(ws, row, "Selling, general & administrative", ASSUMP_ROWS['conn_sga'], sf_conn_rev); row += 1
sf_conn_op = row
ws.cell(row=row, column=1, value="Income from operations").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(HIST_ROWS['conn_op'])}"); c.font = link_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_conn_rev}-{col}{sf_conn_cogs}-{col}{sf_conn_rd}-{col}{sf_conn_sga}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
sf_conn_capex = row = cost_line(ws, row, "Capex", ASSUMP_ROWS['conn_capex'], sf_conn_rev)
c = ws.cell(row=sf_conn_capex, column=2, value=f"={hist_ref(conn_locs['Capex'])}"); c.font = link_font(); c.number_format = CUR0
row += 2

# --- AI (compute-driven) ---
section_row(ws, row, "AI SEGMENT", ncols=2+NF); row += 1
sf_gw = row
ws.cell(row=row, column=1, value="Nameplate compute, period end (GW)").font = lbl_font()
c = ws.cell(row=row, column=2, value=1.4); c.font = input_font(); c.number_format = '0.0'
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['ai_gw'], acol)}")
    cell.font = link_font(); cell.number_format = '0.0'
row += 1
sf_monetizable = row
ws.cell(row=row, column=1, value="  Monetizable share (net of internal Grok training)").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2, value=1.0); c.font = input_font(size=9); c.number_format = PCT1
note(ws, f"B{sf_monetizable}", "Q2'26A is set to 100% because the reported actual revenue already reflects whatever internal usage existed in that quarter — applying a further haircut would double-count it. The forecast quarters apply the 90% assumption from the Assumptions tab.")
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['ai_monetizable'], acol)}")
    cell.font = link_font(size=9); cell.number_format = PCT1
row += 1
sf_rev_gw = row
ws.cell(row=row, column=1, value="  Revenue-generating compute (GW)").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2, value=f"=B{sf_gw}*B{sf_monetizable}"); c.font = formula_font(size=9); c.number_format = '0.0'
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_gw}*{col}{sf_monetizable}")
    cell.font = formula_font(size=9); cell.number_format = '0.0'
row += 1
sf_rev_per_gw = row
ws.cell(row=row, column=1, value="AI infrastructure revenue per revenue-generating GW ($mm/qtr)").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"={hist_ref(ai_locs['AI solutions & infrastructure revenue'])}/B{sf_gw}")
c.font = link_font(); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['ai_rev_per_gw'], acol)}")
    cell.font = link_font(); cell.number_format = CUR0
row += 1
sf_ai_infra_rev = row
ws.cell(row=row, column=1, value="AI infrastructure revenue (revenue-generating GW x $/GW/qtr)").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"={hist_ref(ai_locs['AI solutions & infrastructure revenue'])}"); c.font = link_font(); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_rev_gw}*{col}{sf_rev_per_gw}")
    cell.font = formula_font(); cell.number_format = CUR0
row += 1
sf_grok_rev = row
ws.cell(row=row, column=1, value="Grok subscription revenue  [ASSUMPTION — no historical anchor]").font = lbl_font()
c = ws.cell(row=row, column=2, value=0); c.font = input_font(); c.number_format = CUR0
note(ws, f"B{sf_grok_rev}", "Anchor set to zero because SpaceX does not break Grok subscriptions out in reported actuals — they sit inside the disclosed 'AI solutions and infrastructure' line, which anchors the $/GW rate above. Forecast quarters carry Grok as a separate line so that non-GW-linked revenue is visible; this is a deliberate presentational break from the historical format.")
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['grok_rev'], acol)}")
    cell.font = link_font(); cell.number_format = CUR0
row += 1
sf_cursor_rev = row
ws.cell(row=row, column=1, value="Cursor revenue  [ASSUMPTION — no historical anchor]").font = lbl_font()
c = ws.cell(row=row, column=2, value=0); c.font = input_font(); c.number_format = CUR0
note(ws, f"B{sf_cursor_rev}", "Zero in Q2'26A — Cursor was not yet integrated. Management's $100B ARR trajectory is explicitly stated as 'including contribution from Cursor', so this is partly inorganic revenue that scales with seats rather than gigawatts.")
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['cursor_rev'], acol)}")
    cell.font = link_font(); cell.number_format = CUR0
row += 1
sf_adv_rev = row
ws.cell(row=row, column=1, value="Advertising revenue").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"={hist_ref(ai_locs['Advertising revenue'])}"); c.font = link_font(); c.number_format = CUR0
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={prev_col}{row}*(1+{asum_ref(ASSUMP_ROWS['ai_adv_growth'], acol)})")
    cell.font = formula_font(); cell.number_format = CUR0
    prev_col = col
row += 1
sf_ai_rev = row
ws.cell(row=row, column=1, value="AI revenue").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(HIST_ROWS['ai_rev'])}"); c.font = link_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i,
        value=f"={col}{sf_ai_infra_rev}+{col}{sf_grok_rev}+{col}{sf_cursor_rev}+{col}{sf_adv_rev}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ws.cell(row=row, column=1, value="  QoQ growth %").font = lbl_font(italic=True, size=9)
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"=IFERROR({col}{sf_ai_rev}/{prev_col}{sf_ai_rev}-1,\"n/a\")")
    cell.font = formula_font(size=9); cell.number_format = PCT1
    prev_col = col
row += 1
sf_ai_cogs = row = cost_line(ws, row, "Cost of revenue", ASSUMP_ROWS['ai_cogs'], sf_ai_rev); row += 1
sf_ai_rd = row = cost_line(ws, row, "Research & development", ASSUMP_ROWS['ai_rd'], sf_ai_rev); row += 1
sf_ai_sga = row = cost_line(ws, row, "Selling, general & administrative", ASSUMP_ROWS['ai_sga'], sf_ai_rev); row += 1
sf_ai_op = row
ws.cell(row=row, column=1, value="Income (loss) from operations").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(HIST_ROWS['ai_op'])}"); c.font = link_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_ai_rev}-{col}{sf_ai_cogs}-{col}{sf_ai_rd}-{col}{sf_ai_sga}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
sf_ai_gw_added = row
ws.cell(row=row, column=1, value="  Incremental GW added in quarter").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2, value=0.4); c.font = input_font(size=9); c.number_format = '0.0'
note(ws, f"B{row}", "Q1'26A GW (1.0) to Q2'26A GW (1.4) per KPIs tab")
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_gw}-{prev_col}{sf_gw}")
    cell.font = formula_font(size=9); cell.number_format = '0.0'
    prev_col = col
row += 1
sf_ai_capex_per_gw = row
ws.cell(row=row, column=1, value="  Capex per incremental GW ($mm/GW)").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2, value=f"={hist_ref(ai_locs['Capex'])}/B{sf_ai_gw_added}")
c.font = link_font(size=9); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['ai_capex_per_gw'], acol)}")
    cell.font = link_font(size=9); cell.number_format = CUR0
row += 1
sf_marginal_gw = row
ws.cell(row=row, column=1, value="  Marginal contract rate on new capacity ($mm/GW/qtr)").font = lbl_font(italic=True, size=9)
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['ai_marginal_gw'], acol)}")
    cell.font = link_font(size=9); cell.number_format = CUR0
row += 1
sf_ai_capex = row
ws.cell(row=row, column=1, value="AI Capex (compute + data center + cooling)").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=f"={hist_ref(ai_locs['Capex'])}"); c.font = link_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_ai_gw_added}*{col}{sf_ai_capex_per_gw}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ALLCOLS = ["B"] + SF_FCOLS  # Q2'26A .. Q4'28E, 11 columns total
LAG_QUARTERS = 2  # 6 months

sf_payback_mgmt = row
ws.cell(row=row, column=1, value="  Payback on MANAGEMENT basis (capex/GW / marginal rate, no lag)").font = lbl_font(italic=True, size=9, bold=True)
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_ai_capex_per_gw}/({col}{sf_marginal_gw}*4)")
    cell.font = formula_font(size=9, bold=True); cell.number_format = '0.00" yrs"'
row += 1
sf_payback_mgmt_chk = row
ws.cell(row=row, column=1, value="    Meets the <1-year claim?").font = lbl_font(italic=True, size=9)
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f'=IF({col}{sf_payback_mgmt}<1,"Yes","No")')
    cell.font = formula_font(size=9)
row += 1
sf_payback = row
ws.cell(row=row, column=1, value="  Marginal capex payback (years) — 6mo lag, vintage rate, monetizable GW only").font = lbl_font(italic=True, size=9, bold=True)
for i, col in enumerate(ALLCOLS):
    if i + LAG_QUARTERS < len(ALLCOLS):
        rate_col = ALLCOLS[i + LAG_QUARTERS]
        formula = (f"=0.5+{col}{sf_ai_capex_per_gw}"
                   f"/({rate_col}{sf_rev_per_gw}*{rate_col}{sf_monetizable}*4)")
    else:
        formula = '="n/a"'
    cell = ws.cell(row=row, column=2+i, value=formula)
    cell.font = formula_font(size=9, bold=True); cell.number_format = '0.00" yrs";;;@'
row += 1
sf_payback_check = row
ws.cell(row=row, column=1, value="  Meets SpaceX's <1-year AI compute payback guidance?").font = lbl_font(italic=True, size=9)
for i, col in enumerate(ALLCOLS):
    cell = ws.cell(row=row, column=2+i,
        value=f'=IF({col}{sf_payback}="n/a","n/a",IF({col}{sf_payback}<=1,"Yes","No"))')
    cell.font = formula_font(size=9)
row += 1
ws.cell(row=row, column=1, value="Two payback measures are shown. The MANAGEMENT basis divides capex per GW by the annualised marginal contract rate that new capacity earns, with no build lag and no internal-use haircut — this is the closest reconstruction of what the CFO's 'less than one-year payback' most likely means, and Case 2 reproduces it. The measure below is deliberately stricter. Marginal basis: capex committed in quarter X is assumed to come online and start earning revenue only in quarter X+2 (6-month build/ramp lag — no revenue during that window), monetized at the $/GW rate the model assumes will prevail in quarter X+2 (not a blended, whole-portfolio average). Capex is incurred on ALL GW built, but only the monetizable share earns revenue, so the internal Grok training allocation lengthens payback proportionally. Last 2 forecast columns show \"n/a\" because quarter X+2 falls beyond this model's Q4'28E horizon.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2+NF)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

# --- ORBITAL AI COMPUTE (speculative memo, not integrated into totals) ---
section_row(ws, row, "ORBITAL AI COMPUTE (SPECULATIVE MEMO — additive to terrestrial GW above; excluded from Total Capex/Revenue/Net Income/Payback)", ncols=2+NF); row += 1
sf_orb_gw = row
ws.cell(row=row, column=1, value="Orbital AI compute, period end (GW)").font = lbl_font()
c = ws.cell(row=row, column=2, value=0.0); c.font = input_font(); c.number_format = '0.0'
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['orbital_gw'], acol)}")
    cell.font = link_font(); cell.number_format = '0.0'
row += 1
sf_orb_gw_added = row
ws.cell(row=row, column=1, value="  Incremental orbital GW added in quarter").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2, value=0.0); c.font = input_font(size=9); c.number_format = '0.0'
prev_col = "B"
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_orb_gw}-{prev_col}{sf_orb_gw}")
    cell.font = formula_font(size=9); cell.number_format = '0.0'
    prev_col = col
row += 1
sf_orb_capex_per_gw = row
ws.cell(row=row, column=1, value="  Capex per incremental orbital GW ($mm/GW)").font = lbl_font(italic=True, size=9)
for i, col in enumerate(SF_FCOLS):
    acol = ASSUMP_COLS[i]
    cell = ws.cell(row=row, column=3+i, value=f"={asum_ref(ASSUMP_ROWS['orbital_capex_per_gw'], acol)}")
    cell.font = link_font(size=9); cell.number_format = CUR0
row += 1
sf_orb_capex = row
ws.cell(row=row, column=1, value="Orbital AI capex (satellites + Starship launch) — MEMO, not in Total Capex").font = lbl_font(bold=True)
c = ws.cell(row=row, column=2, value=0); c.font = input_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_orb_gw_added}*{col}{sf_orb_capex_per_gw}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
sf_orb_rev = row
ws.cell(row=row, column=1, value="Orbital AI revenue — MEMO, n/a (no operational satellites, no disclosed monetization plan)").font = lbl_font(italic=True, size=9)
for i, col in enumerate(["B"]+SF_FCOLS):
    cell = ws.cell(row=row, column=2+i, value='="n/a"')
    cell.font = formula_font(size=9)
row += 1
ws.cell(row=row, column=1, value="SpaceX's own stated target: 1 GW of orbital AI compute by end of 2027, 'order of magnitude' annual scaling thereafter (per its S-1 and June 2026 AI1/Gigasat disclosures) — Musk cautioned to 'take this with a grain of salt.' Capex/GW is anchored to the ~$55B total Gigasat program cost against its 1GW target, since there is no comparable-company or operating data for satellite-based AI compute. No revenue line is forecast: no orbital satellites have launched, and SpaceX has not disclosed a monetization plan or timeline. This entire block is excluded from every other calculation in this model (Total Capex, Total AI Revenue, Consolidated P&L, Net Income, Implied Market Cap, AI capex payback) so the core forecast stays anchored to actual reported results — treat these rows as a separate 'blue sky' scenario, not part of the base case.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2+NF)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

# --- TOTAL ---
section_row(ws, row, "TOTAL COMPANY", ncols=2+NF); row += 1
sf_tot_rev = row
ws.cell(row=row, column=1, value="Total revenue").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"=B{sf_space_rev}+B{sf_conn_rev}+B{sf_ai_rev}"); c.font = formula_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_space_rev}+{col}{sf_conn_rev}+{col}{sf_ai_rev}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
sf_tot_op = row
ws.cell(row=row, column=1, value="Total income (loss) from operations").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"=B{sf_space_op}+B{sf_conn_op}+B{sf_ai_op}"); c.font = formula_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_space_op}+{col}{sf_conn_op}+{col}{sf_ai_op}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ws.cell(row=row, column=1, value="  Operating margin %").font = lbl_font(italic=True, size=9)
for i, col in enumerate(["B"]+SF_FCOLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{sf_tot_op}/{col}{sf_tot_rev}")
    cell.font = formula_font(size=9); cell.number_format = PCT1
row += 1
sf_tot_capex = row
ws.cell(row=row, column=1, value="Total Capex").font = formula_font(bold=True)
c = ws.cell(row=row, column=2, value=f"=B{sf_space_capex}+B{sf_conn_capex}+B{sf_ai_capex}"); c.font = formula_font(bold=True); c.number_format = CUR0
for i, col in enumerate(SF_FCOLS):
    cell = ws.cell(row=row, column=3+i, value=f"={col}{sf_space_capex}+{col}{sf_conn_capex}+{col}{sf_ai_capex}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ws.cell(row=row, column=1, value="  of which: AI compute & cooling capex").font = lbl_font(italic=True, size=9)
for i, col in enumerate(["B"]+SF_FCOLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{sf_ai_capex}/{col}{sf_tot_capex}")
    cell.font = formula_font(size=9); cell.number_format = PCT1
row += 1

SF_ROWS = dict(space_rev=sf_space_rev, space_op=sf_space_op, space_capex=sf_space_capex,
               conn_rev=sf_conn_rev, conn_op=sf_conn_op, conn_capex=sf_conn_capex,
               ai_rev=sf_ai_rev, ai_op=sf_ai_op, ai_gw=sf_gw, ai_capex=sf_ai_capex, ai_payback=sf_payback, ai_payback_mgmt=sf_payback_mgmt,
               rev_gw=sf_rev_gw, gw_added=sf_ai_gw_added, grok=sf_grok_rev, cursor=sf_cursor_rev, adv=sf_adv_rev,
               orbital_gw=sf_orb_gw, orbital_capex=sf_orb_capex,
               tot_rev=sf_tot_rev, tot_op=sf_tot_op, tot_capex=sf_tot_capex)

ws.freeze_panes = "C5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Segment Forecast sheet done", SF_ROWS)

# =========================================================================
# Financing & D&A tab layout (fixed row numbers, built later; CP links to these)
# =========================================================================
FIN = "Financing & D&A"
FIN_CAPEX      = 6
FIN_VINTAGE    = 7
FIN_DA         = 8
FIN_OPENCASH   = 11
FIN_INTINC     = 12
FIN_OPENDEBT   = 13
FIN_INTEXP     = 14
FIN_OPINC_PRE  = 15
FIN_NONCASH    = 16
FIN_TAX        = 17
FIN_OCF        = 18
FIN_CAPEX2     = 19
FIN_PREFIN     = 20
FIN_DRAW       = 21
FIN_CLOSECASH  = 22
FIN_CLOSEDEBT  = 23
FIN_TOTDEBT    = 24
FIN_CUMDRAW    = 25
FIN_COLS = ASSUMP_COLS  # B..K = the 10 forecast quarters

# =========================================================================
# 6. CONSOLIDATED P&L (historical + forecast)
# =========================================================================
ws = wb.create_sheet("Consolidated P&L")
sheet_header(ws, "Consolidated P&L — Actuals & Forecast", "$ in millions. Historical columns link to Historical P&L; forecast columns link to Segment Forecast + Assumptions, through Q4 2028")
set_col_widths(ws, [34] + [11]*len(ALLQ) + [3])

row = 4
col_headers(ws, row, ["$mm"] + ALLQ)
row += 1

def cp_seg_row(ws, row, label, hist_row_B, hist_row_C, sf_row):
    """3 historical cols (B,C,D) then NF forecast cols, using CP_COLS mapping."""
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, col in enumerate(CP_COLS):
        if i == 0:
            f = f"='Historical P&L'!B{hist_row_B}"
        elif i == 1:
            f = f"='Historical P&L'!C{hist_row_C}"
        elif i == 2:
            f = f"='Segment Forecast'!B{sf_row}"
        else:
            f = f"='Segment Forecast'!{SF_FCOLS[i-3]}{sf_row}"
        cell = ws.cell(row=row, column=2+i, value=f); cell.font = link_font(); cell.number_format = CUR0
    return row

section_row(ws, row, "REVENUE BY SEGMENT", ncols=1+len(ALLQ)); row += 1
cp_space_rev = row; cp_seg_row(ws, row, "Space", HIST_ROWS['space_rev'], HIST_ROWS['space_rev'], SF_ROWS['space_rev']); row += 1
cp_conn_rev = row; cp_seg_row(ws, row, "Connectivity", HIST_ROWS['conn_rev'], HIST_ROWS['conn_rev'], SF_ROWS['conn_rev']); row += 1
cp_ai_rev = row; cp_seg_row(ws, row, "AI", HIST_ROWS['ai_rev'], HIST_ROWS['ai_rev'], SF_ROWS['ai_rev']); row += 1
cp_tot_rev = row
ws.cell(row=row, column=1, value="Total revenue").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{cp_space_rev}:{col}{cp_ai_rev})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ws.cell(row=row, column=1, value="  YoY growth %").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    if i >= 4:
        prior_col = CP_COLS[i-4]
        cell = ws.cell(row=row, column=2+i, value=f"=IFERROR({col}{cp_tot_rev}/{prior_col}{cp_tot_rev}-1,\"n/a\")")
        cell.font = formula_font(size=9); cell.number_format = PCT1
row += 2

section_row(ws, row, "REVENUE RUN-RATE — MANAGEMENT'S ARR BASIS (exit month x 12)", ncols=1+len(ALLQ)); row += 1
cp_exitmo = row
ws.cell(row=row, column=1, value="Exit-month revenue (final month of quarter)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i,
            value=f"={col}{cp_tot_rev}*Assumptions!$B${ASSUMP_ROWS['dec_share']}")
        cell.font = formula_font()
    cell.number_format = CUR0
row += 1
cp_arr = row
ws.cell(row=row, column=1, value="ARR (exit month x 12)").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_exitmo}*12")
        cell.font = Font(name=FONT, size=11, bold=True); cell.border = top_border
    cell.number_format = CUR0
row += 1
cp_arr_gap = row
ws.cell(row=row, column=1, value="  Gap to management's $100B December 2026 guidance").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2+4, value=f"={CP_COLS[4]}{cp_arr}-100000")
c.font = formula_font(size=9, bold=True); c.number_format = CUR0
note(ws, f"{CP_COLS[4]}{cp_arr_gap}", "CFO Bret Johnsen, Q2'26 call: a trajectory 'including contribution from Cursor, to reach $100 billion of ARR, or annualized revenue run rate by the end of this year, based on our expected revenue in the month of December of this year.'")
row += 1
ws.cell(row=row, column=1, value="Management measures ARR as expected December revenue annualised, not the quarter multiplied by four, and states it inclusive of Cursor. This block puts the model on that same basis so the comparison is like-for-like. The exit-month share is an input on the Assumptions tab.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

section_row(ws, row, "COSTS & EXPENSES", ncols=1+len(ALLQ)); row += 1
cp_cogs = row
ws.cell(row=row, column=1, value="Cost of revenue").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i == 0: f = "=2282"
    elif i == 1: f = "='Historical P&L'!C9+'Historical P&L'!C21+'Historical P&L'!C32"
    elif i == 2: f = "='Historical P&L'!D9+'Historical P&L'!D21+'Historical P&L'!D32"
    else:
        fcol = SF_FCOLS[i-3]
        f = f"='Segment Forecast'!{fcol}{sf_space_cogs}+'Segment Forecast'!{fcol}{sf_conn_cogs}+'Segment Forecast'!{fcol}{sf_ai_cogs}"
    cell = ws.cell(row=row, column=2+i, value=f); cell.font = link_font(); cell.number_format = CUR0
row += 1
cp_rd = row
ws.cell(row=row, column=1, value="Research & development").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i == 0: f = "=1958"
    elif i == 1: f = "='Historical P&L'!C10+'Historical P&L'!C22+'Historical P&L'!C33"
    elif i == 2: f = "='Historical P&L'!D10+'Historical P&L'!D22+'Historical P&L'!D33"
    else:
        fcol = SF_FCOLS[i-3]
        f = f"='Segment Forecast'!{fcol}{sf_space_rd}+'Segment Forecast'!{fcol}{sf_conn_rd}+'Segment Forecast'!{fcol}{sf_ai_rd}"
    cell = ws.cell(row=row, column=2+i, value=f); cell.font = link_font(); cell.number_format = CUR0
row += 1
cp_sga = row
ws.cell(row=row, column=1, value="Selling, general & administrative").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i == 0: f = "=606"
    elif i == 1: f = "='Historical P&L'!C11+'Historical P&L'!C23+'Historical P&L'!C34"
    elif i == 2: f = "='Historical P&L'!D11+'Historical P&L'!D23+'Historical P&L'!D34"
    else:
        fcol = SF_FCOLS[i-3]
        f = f"='Segment Forecast'!{fcol}{sf_space_sga}+'Segment Forecast'!{fcol}{sf_conn_sga}+'Segment Forecast'!{fcol}{sf_ai_sga}"
    cell = ws.cell(row=row, column=2+i, value=f); cell.font = link_font(); cell.number_format = CUR0
row += 1
cp_restruct = row
ws.cell(row=row, column=1, value="Restructuring charges (credits)").font = lbl_font()
restruct_vals = [190, -11, 2] + [0]*NF
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=restruct_vals[i]); cell.font = input_font(); cell.number_format = CUR0
row += 1
cp_impair = row
ws.cell(row=row, column=1, value="Impairment").font = lbl_font()
impair_vals = [5, 0, 0] + [0]*NF
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=impair_vals[i]); cell.font = input_font(); cell.number_format = CUR0
row += 1
cp_tot_costs_pre = row
ws.cell(row=row, column=1, value="Total costs and expenses (excl. new-capex D&A)").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{cp_cogs}:{col}{cp_impair})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
cp_op_income_pre = row
ws.cell(row=row, column=1, value="Operating income (loss) before new-capex D&A").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_tot_rev}-{col}{cp_tot_costs_pre}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0
row += 1
cp_da = row
ws.cell(row=row, column=1, value="Depreciation on new forecast-period capex").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=0)
        cell.font = input_font()
    else:
        fcol = FIN_COLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"='{FIN}'!{fcol}{FIN_DA}")
        cell.font = link_font()
    cell.number_format = CUR0
row += 1
ws.cell(row=row, column=1, value="Historical columns show zero because depreciation on assets already in service at Q2'26 is embedded in the reported cost lines above (and in the cost ratios used to forecast them). Only depreciation on NEW capex spent during the forecast is added here, to avoid double-counting.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 1
cp_tot_costs = row
ws.cell(row=row, column=1, value="Total costs and expenses").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_tot_costs_pre}+{col}{cp_da}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
cp_op_income = row
ws.cell(row=row, column=1, value="Income (loss) from operations").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_tot_rev}-{col}{cp_tot_costs}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ws.cell(row=row, column=1, value="  Operating margin %").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_op_income}/{col}{cp_tot_rev}")
    cell.font = formula_font(size=9); cell.number_format = PCT1
row += 2

section_row(ws, row, "BELOW THE LINE", ncols=1+len(ALLQ)); row += 1
cp_int_exp = row
ws.cell(row=row, column=1, value="Interest expense").font = lbl_font()
hist_int_exp = [-411, -664, -629]
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=hist_int_exp[i]); cell.font = input_font()
    else:
        fcol = FIN_COLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"='{FIN}'!{fcol}{FIN_INTEXP}"); cell.font = link_font()
    cell.number_format = CUR0
row += 1
cp_int_inc = row
ws.cell(row=row, column=1, value="Interest income").font = lbl_font()
hist_int_inc = [98, 213, 340]
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=hist_int_inc[i]); cell.font = input_font()
    else:
        fcol = FIN_COLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"='{FIN}'!{fcol}{FIN_INTINC}"); cell.font = link_font()
    cell.number_format = CUR0
row += 1
cp_other_inc = row
ws.cell(row=row, column=1, value="Other income (expense), net").font = lbl_font()
hist_other = [413, -1876, -86]
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=hist_other[i]); cell.font = input_font()
    else:
        acol = ASSUMP_COLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"={asum_ref(ASSUMP_ROWS['other_inc'], acol)}"); cell.font = link_font()
    cell.number_format = CUR0
row += 1
cp_pretax = row
ws.cell(row=row, column=1, value="Loss before income taxes").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_op_income}+{col}{cp_int_exp}+{col}{cp_int_inc}+{col}{cp_other_inc}")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
cp_tax = row
ws.cell(row=row, column=1, value="Provision for income taxes").font = lbl_font()
hist_tax = [138, 6, 23]
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=hist_tax[i]); cell.font = input_font()
    else:
        acol = ASSUMP_COLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"={asum_ref(ASSUMP_ROWS['tax'], acol)}"); cell.font = link_font()
    cell.number_format = CUR0
row += 1
cp_netloss = row
ws.cell(row=row, column=1, value="Net income (loss)").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_pretax}-{col}{cp_tax}")
    cell.font = Font(name=FONT, size=11, bold=True); cell.number_format = CUR0; cell.border = dbl_top
row += 1
ws.cell(row=row, column=1, value="  Net margin %").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_netloss}/{col}{cp_tot_rev}")
    cell.font = formula_font(size=9); cell.number_format = PCT1
row += 2

section_row(ws, row, "ADJUSTED EBITDA (memo)", ncols=1+len(ALLQ)); row += 1
ws.cell(row=row, column=1, value="Adjusted EBITDA").font = lbl_font(bold=True)
ebitda_hist = [1214, 1127, 3538]
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=ebitda_hist[i]); cell.font = input_font()
    else:
        cell = ws.cell(row=row, column=2+i,
            value=f"={col}{cp_op_income}+{col}{cp_da}+Assumptions!$B${ASSUMP_ROWS['legacy_noncash']}")
        cell.font = formula_font()
    cell.number_format = CUR0
row += 1
ws.cell(row=row, column=1, value="Forecast Adjusted EBITDA = operating income + new-capex depreciation + the legacy non-cash addback held flat from Q2'26A. Now computable because depreciation is an explicit line.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

section_row(ws, row, "IMPLIED MARKET CAP (memo — cloud computing comps)", ncols=1+len(ALLQ)); row += 1
cp_ltm_rev = row
ws.cell(row=row, column=1, value="LTM revenue (trailing 4 quarters)").font = lbl_font()
# Consecutive-quarter columns are C..N (Q1'26A through Q4'28E); LTM needs 4 trailing consecutive quarters,
# so the first computable column is F (Q4'26E) = C+D+E+F. B (Q2'25A) is an isolated prior-year comp and is skipped.
for i, col in enumerate(CP_COLS):
    if i < 4:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        c0, c1_, c2_, c3_ = CP_COLS[i-3], CP_COLS[i-2], CP_COLS[i-1], CP_COLS[i]
        cell = ws.cell(row=row, column=2+i, value=f"=SUM({c0}{cp_tot_rev}:{c3_}{cp_tot_rev})")
        cell.font = formula_font(); cell.number_format = CUR0
row += 1
cp_selmult = row
ws.cell(row=row, column=1, value="Selected Mkt Cap / Revenue multiple (from Assumptions)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i < 4:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i, value=f"=Assumptions!$B${ASSUMP_ROWS['selected_mult']}")
        cell.font = link_font()
    cell.number_format = '0.0"x"'
row += 1
cp_mktcap = row
ws.cell(row=row, column=1, value="Implied market cap ($B)").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(CP_COLS):
    if i < 4:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_ltm_rev}*{col}{cp_selmult}/1000")
        cell.font = Font(name=FONT, size=11, bold=True); cell.border = top_border
    cell.number_format = '$#,##0.0"B"'
row += 1
cp_ltm_ni = row
ws.cell(row=row, column=1, value="LTM net income (loss)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i < 4:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        c0, c1_, c2_, c3_ = CP_COLS[i-3], CP_COLS[i-2], CP_COLS[i-1], CP_COLS[i]
        cell = ws.cell(row=row, column=2+i, value=f"=SUM({c0}{cp_netloss}:{c3_}{cp_netloss})")
        cell.font = formula_font(); cell.number_format = CUR0
row += 1
cp_pe = row
ws.cell(row=row, column=1, value="Implied P/E ratio").font = lbl_font(bold=True)
for i, col in enumerate(CP_COLS):
    if i < 4:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i,
            value=f'=IF({col}{cp_ltm_ni}<=0,"n/m",{col}{cp_mktcap}*1000/{col}{cp_ltm_ni})')
        cell.font = formula_font(bold=True); cell.border = top_border
    cell.number_format = '0.0"x";(0.0"x")'
row += 1
cp_ni_growth = row
ws.cell(row=row, column=1, value="  LTM net income YoY growth % (memo)").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    if i < 8:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        prior_col = CP_COLS[i-4]
        cell = ws.cell(row=row, column=2+i,
            value=f'=IF(OR({prior_col}{cp_ltm_ni}<=0,{col}{cp_ltm_ni}<=0),"n/m",{col}{cp_ltm_ni}/{prior_col}{cp_ltm_ni}-1)')
        cell.font = formula_font(size=9)
    cell.number_format = PCT1
row += 1
cp_peg = row
ws.cell(row=row, column=1, value="Implied PEG ratio").font = lbl_font(bold=True)
for i, col in enumerate(CP_COLS):
    if i < 8:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i,
            value=f'=IF(OR({col}{cp_pe}="n/m",{col}{cp_ni_growth}="n/m",{col}{cp_ni_growth}<=0),"n/m",{col}{cp_pe}/({col}{cp_ni_growth}*100))')
        cell.font = formula_font(bold=True); cell.border = top_border
    cell.number_format = '0.00"x"'
row += 1
ws.cell(row=row, column=1, value="P/E = implied market cap / LTM net income; shows \"n/m\" while LTM net income is <=0. PEG = P/E / LTM net income YoY growth (%, growth expressed as a whole number, e.g. 25 for 25%) — the standard convention, though PEG is designed for steady-state earnings growth and is a rough gauge here given SpaceX's early, lumpy path to profitability.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 1

cp_peg_pe = row
ws.cell(row=row, column=1, value="Implied P/E from peer PEG benchmark (memo)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i < 8:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i,
            value=f'=IF({col}{cp_ni_growth}="n/m","n/m",Assumptions!$B${ASSUMP_ROWS["selected_peg"]}*{col}{cp_ni_growth}*100)')
        cell.font = link_font()
    cell.number_format = '0.0"x"'
row += 1
cp_peg_mktcap = row
ws.cell(row=row, column=1, value="Implied market cap ($B, PEG-based)").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(CP_COLS):
    if i < 8:
        cell = ws.cell(row=row, column=2+i, value=None)
    else:
        cell = ws.cell(row=row, column=2+i,
            value=f'=IF({col}{cp_peg_pe}="n/m","n/m",{col}{cp_peg_pe}*{col}{cp_ltm_ni}/1000)')
        cell.font = Font(name=FONT, size=11, bold=True); cell.border = top_border
    cell.number_format = '$#,##0.0"B";("n/m")'
row += 1
ws.cell(row=row, column=1, value="PEG-based market cap = (selected peer PEG x SpaceX's own LTM net income YoY growth %) x LTM net income. Peer PEG benchmark set: NVIDIA, Palantir, Microsoft (median, editable on Assumptions). Only computable once a prior-year LTM net income base exists (Q4'27E onward) — same constraint as the PEG ratio above. CAUTION: this method is extremely sensitive to SpaceX's near-term earnings growth rate, which is elevated (100-750%+ YoY) because net income is compounding off a very small base shortly after turning profitable — peer PEGs assume much more moderate steady-state growth (10-40%/yr). The result is implausibly large implied values below; this line is shown for completeness against the PEG methodology the user requested, not as a reliable valuation. The revenue-multiple market cap above is the more defensible estimate at this stage.").font = lbl_font(italic=True, size=8, color="C00000")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 1
ws.cell(row=row, column=1, value="Note: comparable-multiple valuation only; ignores SpaceX's actual IPO-day market cap, dual-class share structure, and founder-control premium/discount. Not a price target.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

note(ws, f"B{cp_tot_rev}", "Historical actuals: SpaceX Q2 2026 earnings release. Forecast: linked from Segment Forecast tab, driven by Assumptions tab.")
ws.freeze_panes = "B5"

# --- CAPITAL EXPENDITURES ---
section_row(ws, row, "CAPITAL EXPENDITURES (memo)", ncols=1+len(ALLQ)); row += 1
cp_space_capex = row; cp_seg_row(ws, row, "Space capex", space_locs['Capex'], space_locs['Capex'], SF_ROWS['space_capex']); row += 1
cp_conn_capex = row; cp_seg_row(ws, row, "Connectivity capex", conn_locs['Capex'], conn_locs['Capex'], SF_ROWS['conn_capex']); row += 1
cp_ai_capex = row; cp_seg_row(ws, row, "AI capex (compute + data center + cooling)", ai_locs['Capex'], ai_locs['Capex'], SF_ROWS['ai_capex']); row += 1
cp_tot_capex = row
ws.cell(row=row, column=1, value="Total capex").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{cp_space_capex}:{col}{cp_ai_capex})")
    cell.font = Font(name=FONT, size=11, bold=True); cell.number_format = CUR0; cell.border = dbl_top
row += 1
ws.cell(row=row, column=1, value="  of which: AI compute & cooling %").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_ai_capex}/{col}{cp_tot_capex}")
    cell.font = formula_font(size=9); cell.number_format = PCT1
row += 1
cp_cum_capex = row
ws.cell(row=row, column=1, value="Cumulative AI capex, Q3'26E onward").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=None)
    elif i == 3:
        cell = ws.cell(row=row, column=2+i, value=f"={col}{cp_ai_capex}")
    else:
        prior_col = CP_COLS[i-1]
        cell = ws.cell(row=row, column=2+i, value=f"={prior_col}{cp_cum_capex}+{col}{cp_ai_capex}")
    cell.font = formula_font(size=9); cell.number_format = CUR0
row += 1
ws.cell(row=row, column=1, value="AI capex per incremental GW is a blended compute + data-center + cooling + power build cost, editable on Assumptions. Space/Connectivity capex are modeled as % of segment revenue.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(ALLQ))
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

CP_ROWS = dict(space_rev=cp_space_rev, conn_rev=cp_conn_rev, ai_rev=cp_ai_rev, tot_rev=cp_tot_rev, op_income=cp_op_income, netloss=cp_netloss,
               ltm_rev=cp_ltm_rev, mktcap=cp_mktcap, pe=cp_pe, peg=cp_peg, peg_mktcap=cp_peg_mktcap,
               op_income_pre=cp_op_income_pre, da=cp_da, arr=cp_arr,
               space_capex=cp_space_capex, conn_capex=cp_conn_capex, ai_capex=cp_ai_capex, tot_capex=cp_tot_capex)
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Consolidated P&L sheet done", CP_ROWS)

# =========================================================================
# 6b. FINANCING & D&A SCHEDULE
# =========================================================================
ws = wb.create_sheet(FIN)
sheet_header(ws, "Depreciation & Financing Schedule", "$ in millions. Forecast quarters only. Drives the depreciation and interest lines on the Consolidated P&L.")
set_col_widths(ws, [44] + [11]*NF + [3])

col_headers(ws, 4, ["$mm"] + FQ)

A = lambda r: f"Assumptions!$B${r}"
CPF = lambda i: CP_COLS[3+i]   # Consolidated P&L column for forecast quarter i

section_row(ws, 5, "DEPRECIATION ON NEW FORECAST-PERIOD CAPEX", ncols=1+NF)

ws.cell(row=FIN_CAPEX, column=1, value="Total capex in quarter (vintage)").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_CAPEX, column=2+i, value=f"='Consolidated P&L'!{CPF(i)}{CP_ROWS['tot_capex']}")
    c.font = link_font(); c.number_format = CUR0

ws.cell(row=FIN_VINTAGE, column=1, value="  Quarterly depreciation from this vintage").font = lbl_font(italic=True, size=9)
for i, col in enumerate(FIN_COLS):
    f = (f"={col}{FIN_CAPEX}*({A(ASSUMP_ROWS['it_share'])}/{A(ASSUMP_ROWS['it_life'])}"
         f"+(1-{A(ASSUMP_ROWS['it_share'])})/{A(ASSUMP_ROWS['fac_life'])})")
    c = ws.cell(row=FIN_VINTAGE, column=2+i, value=f)
    c.font = formula_font(size=9); c.number_format = CUR0
note(ws, f"B{FIN_VINTAGE}", "Straight-line: IT share over IT life, facility share over facility life. No vintage reaches the end of its life inside this forecast window, so depreciation accumulates without retirements.")

ws.cell(row=FIN_DA, column=1, value="Depreciation on new capex (cumulative, in service)").font = formula_font(bold=True)
for i, col in enumerate(FIN_COLS):
    lag_i = i - 2   # da_lag = 2 quarters
    if lag_i < 0:
        f = "=0"
    elif i == 0:
        f = "=0"
    else:
        prev = FIN_COLS[i-1]
        f = f"={prev}{FIN_DA}+{FIN_COLS[lag_i]}{FIN_VINTAGE}"
    c = ws.cell(row=FIN_DA, column=2+i, value=f)
    c.font = formula_font(bold=True); c.number_format = CUR0; c.border = top_border
note(ws, f"B{FIN_DA}", "Each quarter adds the depreciation increment from the vintage spent 2 quarters earlier (assets are not depreciated until placed in service), and carries forward all vintages already in service.")

section_row(ws, 10, "CASH, DEBT & INTEREST", ncols=1+NF)

ws.cell(row=FIN_OPENCASH, column=1, value="Opening cash & securities").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    f = f"={A(ASSUMP_ROWS['opening_cash'])}" if i == 0 else f"={FIN_COLS[i-1]}{FIN_CLOSECASH}"
    c = ws.cell(row=FIN_OPENCASH, column=2+i, value=f); c.font = formula_font(); c.number_format = CUR0

ws.cell(row=FIN_INTINC, column=1, value="Interest income (opening cash x yield)").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_INTINC, column=2+i, value=f"={col}{FIN_OPENCASH}*{A(ASSUMP_ROWS['cash_yield'])}/4")
    c.font = formula_font(); c.number_format = CUR0

ws.cell(row=FIN_OPENDEBT, column=1, value="Opening NEW debt balance").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    f = "=0" if i == 0 else f"={FIN_COLS[i-1]}{FIN_CLOSEDEBT}"
    c = ws.cell(row=FIN_OPENDEBT, column=2+i, value=f); c.font = formula_font(); c.number_format = CUR0

ws.cell(row=FIN_INTEXP, column=1, value="Interest expense (existing + new debt)").font = formula_font(bold=True)
for i, col in enumerate(FIN_COLS):
    f = (f"=-({A(ASSUMP_ROWS['exist_debt'])}*{A(ASSUMP_ROWS['exist_rate'])}/4"
         f"+{col}{FIN_OPENDEBT}*{A(ASSUMP_ROWS['new_rate'])}/4)")
    c = ws.cell(row=FIN_INTEXP, column=2+i, value=f)
    c.font = formula_font(bold=True); c.number_format = CUR0
note(ws, f"B{FIN_INTEXP}", "Charged on the OPENING debt balance, so this quarter's new draw does not feed back into this quarter's interest — that keeps the schedule free of circular references.")

ws.cell(row=FIN_OPINC_PRE, column=1, value="Operating income before new-capex D&A").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_OPINC_PRE, column=2+i, value=f"='Consolidated P&L'!{CPF(i)}{CP_ROWS['op_income_pre']}")
    c.font = link_font(); c.number_format = CUR0

ws.cell(row=FIN_NONCASH, column=1, value="Add back: legacy non-cash charges").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_NONCASH, column=2+i, value=f"={A(ASSUMP_ROWS['legacy_noncash'])}")
    c.font = link_font(); c.number_format = CUR0

ws.cell(row=FIN_TAX, column=1, value="Less: cash taxes").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_TAX, column=2+i, value=f"=-{asum_ref(ASSUMP_ROWS['tax'], ASSUMP_COLS[i])}")
    c.font = link_font(); c.number_format = CUR0

ws.cell(row=FIN_OCF, column=1, value="Operating cash flow (proxy)").font = formula_font(bold=True)
for i, col in enumerate(FIN_COLS):
    f = f"={col}{FIN_OPINC_PRE}+{col}{FIN_NONCASH}+{col}{FIN_TAX}+{col}{FIN_INTINC}+{col}{FIN_INTEXP}"
    c = ws.cell(row=FIN_OCF, column=2+i, value=f)
    c.font = formula_font(bold=True); c.number_format = CUR0; c.border = top_border
note(ws, f"B{FIN_OCF}", "Proxy only: excludes working capital movements and deferred revenue, which have been a meaningful source of cash for SpaceX historically. Treat the funding gap below as indicative of scale, not a precise cash forecast.")

ws.cell(row=FIN_CAPEX2, column=1, value="Less: capex").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_CAPEX2, column=2+i, value=f"=-{col}{FIN_CAPEX}")
    c.font = formula_font(); c.number_format = CUR0

ws.cell(row=FIN_PREFIN, column=1, value="Cash before financing").font = formula_font(bold=True)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_PREFIN, column=2+i,
        value=f"={col}{FIN_OPENCASH}+{col}{FIN_OCF}+{col}{FIN_CAPEX2}")
    c.font = formula_font(bold=True); c.number_format = CUR0; c.border = top_border

ws.cell(row=FIN_DRAW, column=1, value="New debt drawn to hold minimum cash").font = formula_font(bold=True)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_DRAW, column=2+i,
        value=f"=MAX(0,{A(ASSUMP_ROWS['min_cash'])}-{col}{FIN_PREFIN})")
    c.font = formula_font(bold=True); c.number_format = CUR0

ws.cell(row=FIN_CLOSECASH, column=1, value="Closing cash & securities").font = formula_font(bold=True)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_CLOSECASH, column=2+i, value=f"={col}{FIN_PREFIN}+{col}{FIN_DRAW}")
    c.font = formula_font(bold=True); c.number_format = CUR0

ws.cell(row=FIN_CLOSEDEBT, column=1, value="Closing NEW debt balance").font = formula_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_CLOSEDEBT, column=2+i, value=f"={col}{FIN_OPENDEBT}+{col}{FIN_DRAW}")
    c.font = formula_font(); c.number_format = CUR0

ws.cell(row=FIN_TOTDEBT, column=1, value="Total debt outstanding (existing + new)").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_TOTDEBT, column=2+i,
        value=f"={A(ASSUMP_ROWS['exist_debt'])}+{col}{FIN_CLOSEDEBT}")
    c.font = Font(name=FONT, size=11, bold=True); c.number_format = CUR0; c.border = dbl_top

ws.cell(row=FIN_CUMDRAW, column=1, value="  Memo: cumulative new debt raised").font = lbl_font(italic=True, size=9)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=FIN_CUMDRAW, column=2+i, value=f"={col}{FIN_CLOSEDEBT}")
    c.font = formula_font(size=9); c.number_format = CUR0

# ---- Free cash flow statement, coverage checks and DCF ----
r = FIN_CUMDRAW + 2
section_row(ws, r, "QUARTERLY FREE CASH FLOW", ncols=1+NF); r += 1
fcf_ebitda = r
ws.cell(row=r, column=1, value="EBITDA (op income before new-capex D&A + legacy non-cash)").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"={col}{FIN_OPINC_PRE}+{col}{FIN_NONCASH}")
    c.font = formula_font(); c.number_format = CUR0
r += 1
fcf_tax = r
ws.cell(row=r, column=1, value="Less: cash taxes").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"={col}{FIN_TAX}"); c.font = link_font(); c.number_format = CUR0
r += 1
fcf_capex = r
ws.cell(row=r, column=1, value="Less: capex").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"={col}{FIN_CAPEX2}"); c.font = link_font(); c.number_format = CUR0
r += 1
fcf_ufcf = r
ws.cell(row=r, column=1, value="UNLEVERED FREE CASH FLOW").font = Font(name=FONT, size=11, bold=True)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"={col}{fcf_ebitda}+{col}{fcf_tax}+{col}{fcf_capex}")
    c.font = Font(name=FONT, size=11, bold=True); c.number_format = CUR0; c.border = dbl_top
r += 1
fcf_lev = r
ws.cell(row=r, column=1, value="Levered free cash flow (after interest)").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"={col}{fcf_ufcf}+{col}{FIN_INTEXP}+{col}{FIN_INTINC}")
    c.font = formula_font(); c.number_format = CUR0
r += 1
ws.cell(row=r, column=1, value="Excludes working capital and deferred revenue movements, which have historically been a source of cash for SpaceX. Unlevered FCF is the input to the DCF below.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1+NF)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
r += 2

section_row(ws, r, "CREDIT CHECKS", ncols=1+NF); r += 1
cc_cov = r
ws.cell(row=r, column=1, value="EBITDA / interest expense (coverage, x)").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"=IFERROR({col}{fcf_ebitda}/-{col}{FIN_INTEXP},\"n/m\")")
    c.font = formula_font(); c.number_format = '0.0"x"'
r += 1
cc_lev = r
ws.cell(row=r, column=1, value="Total debt / annualised EBITDA (leverage, x)").font = lbl_font()
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"=IFERROR({col}{FIN_TOTDEBT}/({col}{fcf_ebitda}*4),\"n/m\")")
    c.font = formula_font(); c.number_format = '0.0"x"'
r += 1
cc_flag = r
ws.cell(row=r, column=1, value="  Leverage within a 6.0x investment-grade-ish ceiling?").font = lbl_font(italic=True, size=9)
for i, col in enumerate(FIN_COLS):
    c = ws.cell(row=r, column=2+i, value=f"=IF(ISNUMBER({col}{cc_lev}),IF({col}{cc_lev}<=6,\"Yes\",\"NO\"),\"n/m\")")
    c.font = formula_font(size=9)
r += 1
ws.cell(row=r, column=1, value="A 6.0x net-debt/EBITDA ceiling is a rough high-yield covenant marker, not a rating agency threshold. Where this reads NO, the capex plan is not fundable with debt alone on these earnings and would require equity or partner/SPV structures of the kind CoreWeave uses.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1+NF)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
r += 2

section_row(ws, r, "DCF (read the terminal-value warning below before using this)", ncols=1+NF); r += 1
dcf_wacc = r
ws.cell(row=r, column=1, value="WACC (annual)").font = lbl_font()
c = ws.cell(row=r, column=2, value=0.115); c.font = input_font(); c.number_format = PCT1; c.fill = yellow_fill; c.border = box
r += 1
dcf_g = r
ws.cell(row=r, column=1, value="Terminal growth rate").font = lbl_font()
c = ws.cell(row=r, column=2, value=0.04); c.font = input_font(); c.number_format = PCT1; c.fill = yellow_fill; c.border = box
r += 1
dcf_tfcf = r
ws.cell(row=r, column=1, value="Normalised terminal-year FCF ($mm/yr)").font = lbl_font()
c = ws.cell(row=r, column=2, value=40000); c.font = input_font(); c.number_format = CUR0; c.fill = yellow_fill; c.border = box
note(ws, f"B{dcf_tfcf}", "Set MANUALLY, not carried from the forecast. Q4'28E FCF is deeply negative because the build-out is still mid-flight, so extrapolating it would value the company below zero. This cell asks what steady-state FCF looks like once capex normalises toward depreciation — an assumption the model cannot derive.")
r += 1
dcf_pv = r
ws.cell(row=r, column=1, value="PV of forecast unlevered FCF (Q3'26E-Q4'28E)").font = lbl_font()
terms = "+".join([f"{col}{fcf_ufcf}/(1+$B${dcf_wacc}/4)^{i+1}" for i, col in enumerate(FIN_COLS)])
c = ws.cell(row=r, column=2, value=f"={terms}"); c.font = formula_font(); c.number_format = CUR0; c.border = box
r += 1
dcf_tv = r
ws.cell(row=r, column=1, value="PV of terminal value").font = lbl_font()
c = ws.cell(row=r, column=2,
    value=f"=$B${dcf_tfcf}*(1+$B${dcf_g})/($B${dcf_wacc}-$B${dcf_g})/(1+$B${dcf_wacc})^2.5")
c.font = formula_font(); c.number_format = CUR0; c.border = box
r += 1
dcf_ev = r
ws.cell(row=r, column=1, value="Enterprise value").font = formula_font(bold=True)
c = ws.cell(row=r, column=2, value=f"=B{dcf_pv}+B{dcf_tv}"); c.font = formula_font(bold=True); c.number_format = CUR0; c.border = top_border
r += 1
dcf_eq = r
ws.cell(row=r, column=1, value="Equity value (EV less closing net debt)").font = Font(name=FONT, size=11, bold=True)
c = ws.cell(row=r, column=2,
    value=f"=B{dcf_ev}-({FIN_COLS[-1]}{FIN_TOTDEBT}-{FIN_COLS[-1]}{FIN_CLOSECASH})")
c.font = Font(name=FONT, size=11, bold=True); c.number_format = CUR0; c.border = dbl_top
r += 1
dcf_tvshare = r
ws.cell(row=r, column=1, value="  Terminal value as % of enterprise value").font = lbl_font(italic=True, size=9)
c = ws.cell(row=r, column=2, value=f"=B{dcf_tv}/B{dcf_ev}"); c.font = formula_font(size=9, bold=True); c.number_format = PCT1
r += 1
ws.cell(row=r, column=1, value="WARNING — READ BEFORE CITING THIS NUMBER. Forecast-period FCF is negative throughout, so the entire equity value rests on the terminal assumption, and the terminal-value share above will typically read well over 100% (the forecast period subtracts value). That makes this a restatement of the manually-entered normalised FCF cell, not an independent valuation. It is included because a cash-flow-based cross-check was requested, and it is more honest than the revenue-multiple and PEG blocks on the Consolidated P&L tab only in that its fragility is explicit. For a company mid-build with negative FCF, no DCF is informative: use it to test what steady-state cash flow would justify a given valuation, by changing the normalised FCF cell until equity value matches the market.").font = lbl_font(italic=True, size=8, color="C00000")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1+NF)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
ws.row_dimensions[r].height = 58
r += 2

ws.cell(row=r, column=1, value="This schedule closes two gaps flagged in the sanity check. (1) Depreciation: the forecast previously carried no charge for the new capex being spent, so operating income was overstated; new-capex depreciation is now an explicit Consolidated P&L line. (2) Financing: interest expense was previously a flat assumption unrelated to the debt implied by the capex plan. Debt is now drawn whenever cash would fall below the minimum buffer, and interest is charged on the actual balance. Both interest expense and interest income on the Consolidated P&L now link here. Simplifications: no working capital or deferred revenue movements, no debt repayment or refinancing, no equity issuance, and no asset retirements inside the window.").font = lbl_font(size=8, italic=True, color="595959")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1+NF)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
ws.row_dimensions[r].height = 60

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Financing & D&A sheet done")

# =========================================================================
# 7. BALANCE SHEET & CASH FLOW (actuals)
# =========================================================================
ws = wb.create_sheet("Balance Sheet & CF")
sheet_header(ws, "Balance Sheet & Cash Flow (Actuals)", "$ in millions. Source: SpaceX Q2 2026 earnings release, Consolidated Balance Sheets and Selected Cash Flow Information")
set_col_widths(ws, [40, 14, 14, 3])

row = 4
col_headers(ws, row, ["Balance Sheet, $mm", "Jun 30, 2026", "Dec 31, 2025"])
row += 1
section_row(ws, row, "ASSETS", ncols=3); row += 1
bs_ca_start = row
for label, a, b in [
    ("Cash and cash equivalents", 93522, 24747),
    ("Marketable securities", 6487, 0),
    ("Accounts receivable, net", 3596, 1579),
    ("Inventory", 2718, 2416),
    ("Prepaid expenses and other current assets", 1724, 2210),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
    row += 1
bs_ca_end = row - 1
tca_row = row
ws.cell(row=row, column=1, value="Total current assets").font = formula_font(bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{bs_ca_start}:{col}{bs_ca_end})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
nc_start = row
for label, a, b in [
    ("Property, plant, and equipment, net", 65736, 42602),
    ("Finance lease right-of-use assets", 1118, 1260),
    ("Intangible assets, net", 1318, 1548),
    ("Digital assets", 1098, 1637),
    ("Goodwill", 11645, 11809),
    ("Deferred tax assets", 354, 141),
    ("Other assets", 3454, 2130),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
    row += 1
nc_end = row - 1
tot_assets_row = row
ws.cell(row=row, column=1, value="Total assets").font = Font(name=FONT, bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{tca_row}+SUM({col}{nc_start}:{col}{nc_end})")
    cell.font = Font(name=FONT, bold=True); cell.number_format = CUR0; cell.border = dbl_top
row += 2

section_row(ws, row, "LIABILITIES", ncols=3); row += 1
lc_start = row
for label, a, b in [
    ("Accounts payable", 8243, 11792),
    ("Deferred revenue, current", 7977, 6111),
    ("Debt and finance leases, current", 2525, 928),
    ("Accrued expenses and other current liabilities", 2377, 2569),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
    row += 1
lc_end = row - 1
tot_cl_row = row
ws.cell(row=row, column=1, value="Total current liabilities").font = formula_font(bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{lc_start}:{col}{lc_end})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
ll_start = row
for label, a, b in [
    ("Deferred revenue, net of current", 6309, 6005),
    ("Debt and finance leases, net of current", 36839, 21968),
    ("Other liabilities", 1276, 1381),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
    row += 1
ll_end = row - 1
tot_liab_row = row
ws.cell(row=row, column=1, value="Total liabilities").font = Font(name=FONT, bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{tot_cl_row}+SUM({col}{ll_start}:{col}{ll_end})")
    cell.font = Font(name=FONT, bold=True); cell.number_format = CUR0; cell.border = top_border
row += 2

section_row(ws, row, "REDEEMABLE CONVERTIBLE PREFERRED STOCK & EQUITY", ncols=3); row += 1
pref_row = row
ws.cell(row=row, column=1, value="Redeemable convertible preferred stock").font = lbl_font()
for i, v in enumerate([0, 38752]):
    cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
row += 1
equity_row = row
ws.cell(row=row, column=1, value="Total shareholders' equity").font = lbl_font()
for i, v in enumerate([127224, 2573]):
    cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
row += 1
tot_liab_eq_row = row
ws.cell(row=row, column=1, value="Total liabilities, preferred stock & equity").font = Font(name=FONT, bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{tot_liab_row}+{col}{pref_row}+{col}{equity_row}")
    cell.font = Font(name=FONT, bold=True); cell.number_format = CUR0; cell.border = dbl_top
row += 1
ws.cell(row=row, column=1, value="Balance check (Assets - Liab&Equity)").font = lbl_font(italic=True, size=9)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{tot_assets_row}-{col}{tot_liab_eq_row}")
    cell.font = formula_font(size=9); cell.number_format = CUR0

row += 3
col_headers(ws, row, ["Cash Flow (6 months ended June 30), $mm", "2026", "2025"])
row += 1
cf_start = row
for label, a, b in [
    ("Net cash provided by operating activities", 3466, 351),
    ("Net cash used in investing activities", -34487, -6032),
    ("Net cash provided by financing activities", 100291, 9199),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate([a, b]):
        cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0; cell.border = box
    row += 1
cf_end = row - 1
change_row = row
ws.cell(row=row, column=1, value="Net change in cash & restricted cash").font = formula_font(bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"=SUM({col}{cf_start}:{col}{cf_end})")
    cell.font = formula_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
begin_row = row
ws.cell(row=row, column=1, value="Cash & restricted cash, beginning of period").font = lbl_font()
for i, v in enumerate([25124, 11501]):
    cell = ws.cell(row=row, column=2+i, value=v); cell.font = input_font(); cell.number_format = CUR0
row += 1
ws.cell(row=row, column=1, value="Cash & restricted cash, end of period").font = Font(name=FONT, bold=True)
for i, col in enumerate(["B", "C"]):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{change_row}+{col}{begin_row}")
    cell.font = Font(name=FONT, bold=True); cell.number_format = CUR0; cell.border = top_border
row += 2

ws.cell(row=row, column=1, value="Key financing events: IPO closed June 12, 2026 (638.9mm Class A shares, ~$85.7B net proceeds); $25B senior notes issued June 26, 2026 (5 tranches, 2031-2056, weighted avg rate 5.855%).").font = lbl_font(size=9, italic=True, color="595959")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Balance Sheet & CF sheet done")

# =========================================================================
# 7b. TERRESTRIAL vs ORBITAL CAPEX PER GW
# =========================================================================
ws = wb.create_sheet("Terrestrial vs Orbital")
sheet_header(ws, "Capex per GW of AI Compute: Terrestrial vs Orbital",
    "$ in millions per GW of AVERAGE compute power. Bottom-up build from SpaceX's disclosed AI1 specifications and published Starship launch economics. Memo/analysis tab — does not feed the forecast.")
set_col_widths(ws, [46, 15, 15, 3, 78])

NOTE_COL = 5
def nrow(ws, row, text):
    c = ws.cell(row=row, column=NOTE_COL, value=text)
    c.font = lbl_font(size=8, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")

row = 4
col_headers(ws, row, ["$mm per GW (average compute)", "Terrestrial", "Orbital"])
ws.cell(row=row, column=NOTE_COL, value="Basis / source").font = hdr_font(size=10, color="000000")
ws.cell(row=row, column=NOTE_COL).fill = grey_fill
ws.cell(row=row, column=NOTE_COL).border = box
row += 1

# ---------------- INPUTS ----------------
section_row(ws, row, "PHYSICAL INPUTS (editable)", ncols=3); row += 1

def two_col_input(ws, row, label, tval, oval, fmt, note_txt):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for j, v in enumerate([tval, oval]):
        c = ws.cell(row=row, column=2+j, value=v)
        if v is None:
            c.value = "n/a"; c.font = lbl_font(size=9, color="808080")
        else:
            c.font = input_font(); c.number_format = fmt; c.fill = yellow_fill
        c.border = box
    nrow(ws, row, note_txt)
    return row

r_avg_kw   = two_col_input(ws, row, "Average compute power per satellite (kW)", None, 120, NUM0,
    "SpaceX AI1 disclosure (Jun 8, 2026): 150 kW peak, 120 kW average compute payload. Average is the right basis since revenue tracks delivered compute."); row += 1
r_kw_ton   = two_col_input(ws, row, "Compute power density (kW per ton)", None, 70, NUM0,
    "AI1 disclosed spec: 70 kW per ton. Musk's forward-looking claim of 100 kW/ton is tested in the sensitivity grid below."); row += 1
r_launch   = two_col_input(ws, row, "Launch cost ($ per kg to LEO)", None, 200, '$#,##0',
    "Starship target range is roughly $100-500/kg once fully reusable; Musk's aspiration is ~$10/kg. Falcon 9 today is ~$2,700-3,000/kg. $200/kg used as the base case."); row += 1
r_solar_w  = two_col_input(ws, row, "Solar array cost ($ per W)", None, 1.50, '$#,##0.00',
    "AI1 carries a 150 kW array at 250 W/m², using SpaceX-manufactured cells from its Bastrop facility. Assumes mass-production economics well below traditional space-solar costs."); row += 1
r_rad_m2   = two_col_input(ws, row, "Radiator cost ($ per m2)", None, 4000, '$#,##0',
    "AI1 carries a 110 m² deployable liquid radiator with redundant pumping loops and micrometeoroid shielding."); row += 1
r_rad_area = two_col_input(ws, row, "Radiator area per satellite (m2)", None, 110, NUM0,
    "AI1 disclosed spec."); row += 1
r_bus      = two_col_input(ws, row, "Satellite bus / structure / integration ($mm each)", None, 1.2, '$#,##0.00',
    "Structure, avionics, propulsion, deployment mechanisms and assembly, excluding compute, solar and radiator. Benchmarked up from Starlink V2 unit economics for a larger, more complex spacecraft."); row += 1
r_premium  = two_col_input(ws, row, "Space-qualification premium on compute hardware", None, 1.30, '0.00"x"',
    "Radiation tolerance, thermal packaging and lower production volumes versus commodity data-centre racks."); row += 1
r_compute  = two_col_input(ws, row, "Compute hardware cost, terrestrial basis ($mm per GW)", 23000, None, CUR0,
    "Epoch AI puts servers/GPUs at ~60% of a $38B/GW build; Bernstein's rack build-up ($3.4M compute of a $5.9M GB200 NVL72 rack) implies ~$24B/GW. Orbital applies the premium above to this same figure."); row += 1
r_bldg     = two_col_input(ws, row, "Building, land & site works ($mm per GW)", 4000, 0, CUR0,
    "Eliminated in orbit — no land, no shell, no zoning."); row += 1
r_power    = two_col_input(ws, row, "Grid power delivery & backup ($mm per GW)", 6500, 0, CUR0,
    "Substation, transformers, UPS and backup generation. Eliminated in orbit: the solar array below replaces it, and sunlight is free."); row += 1
r_cool     = two_col_input(ws, row, "Cooling plant & water systems ($mm per GW)", 4500, 0, CUR0,
    "Chillers, CRAHs, liquid loops and water infrastructure. Eliminated in orbit: replaced by the radiator line below, though vacuum makes heat rejection harder, not easier, per unit area."); row += 1
r_ground   = two_col_input(ws, row, "Ground segment / optical links ($mm per GW)", 0, 1500, CUR0,
    "Orbital needs laser inter-satellite links and ground gateways to move data; terrestrial fibre is assumed included in site works."); row += 1

row += 1
# ---------------- DERIVED ----------------
section_row(ws, row, "DERIVED PHYSICAL QUANTITIES (orbital)", ncols=3); row += 1
r_sats = row
ws.cell(row=row, column=1, value="Satellites required per GW").font = lbl_font()
ws.cell(row=row, column=2, value="n/a").font = lbl_font(size=9, color="808080")
c = ws.cell(row=row, column=3, value=f"=1000000/C{r_avg_kw}"); c.font = formula_font(); c.number_format = NUM0
nrow(ws, row, "1 GW of average compute divided by 120 kW per satellite. Musk noted one AI1 is roughly equivalent to a single NVIDIA GB300 rack.")
row += 1
r_mass = row
ws.cell(row=row, column=1, value="Mass to orbit per GW (metric tons)").font = lbl_font()
ws.cell(row=row, column=2, value="n/a").font = lbl_font(size=9, color="808080")
c = ws.cell(row=row, column=3, value=f"=1000000/C{r_kw_ton}"); c.font = formula_font(); c.number_format = NUM0
nrow(ws, row, "1 GW divided by 70 kW/ton. For scale, this is well above SpaceX's total mass launched to date in a year — the build-out is launch-cadence constrained, not just capital constrained.")
row += 1

row += 1
# ---------------- CAPEX BUILD ----------------
section_row(ws, row, "CAPEX BUILD-UP PER GW", ncols=3); row += 1

def build_line(ws, row, label, tformula, oformula, note_txt):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for j, f in enumerate([tformula, oformula]):
        c = ws.cell(row=row, column=2+j, value=f)
        c.font = formula_font(); c.number_format = CUR0; c.border = box
    nrow(ws, row, note_txt)
    return row

b_compute = build_line(ws, row, "Compute hardware",
    f"=B{r_compute}", f"=B{r_compute}*C{r_premium}",
    "The dominant cost on both sides, and the one orbital cannot avoid."); row += 1
b_bldg = build_line(ws, row, "Building, land & site works", f"=B{r_bldg}", f"=C{r_bldg}",
    "Orbital saving."); row += 1
b_power = build_line(ws, row, "Grid power delivery & backup", f"=B{r_power}", f"=C{r_power}",
    "Orbital saving — this is the 'free energy' advantage in capex terms."); row += 1
b_cool = build_line(ws, row, "Cooling plant & water systems", f"=B{r_cool}", f"=C{r_cool}",
    "Orbital saving on paper, but see the radiator line — heat rejection is replaced, not removed."); row += 1
b_solar = build_line(ws, row, "Solar array (replaces grid power)", "=0",
    f"=C{r_sats}*150*1000*C{r_solar_w}/1000000",
    "150 kW array per satellite at the $/W input. This is the capex you pay to get 'free' energy."); row += 1
b_rad = build_line(ws, row, "Deployable radiators (replace cooling)", "=0",
    f"=C{r_sats}*C{r_rad_area}*C{r_rad_m2}/1000000",
    "110 m² per satellite. Vacuum removes water and chiller cost but radiative heat rejection needs large deployable area — several independent analyses argue AI1's disclosed radiator area is the binding physical constraint on the design."); row += 1
b_bus = build_line(ws, row, "Satellite bus, structure & integration", "=0",
    f"=C{r_sats}*C{r_bus}",
    "No terrestrial analogue — this is pure incremental cost for orbital."); row += 1
b_launch = build_line(ws, row, "Launch to LEO", "=0",
    f"=C{r_mass}*C{r_launch}/1000",
    "Mass per GW times $/kg. Notably NOT the dominant term at Starship target pricing."); row += 1
b_ground = build_line(ws, row, "Ground segment / optical links", f"=B{r_ground}", f"=C{r_ground}",
    "Orbital-only cost."); row += 1

b_total = row
ws.cell(row=row, column=1, value="TOTAL CAPEX PER GW").font = Font(name=FONT, size=11, bold=True)
for j, col in enumerate(["B", "C"]):
    c = ws.cell(row=row, column=2+j, value=f"=SUM({col}{b_compute}:{col}{b_ground})")
    c.font = Font(name=FONT, size=11, bold=True); c.number_format = CUR0; c.border = dbl_top
nrow(ws, row, "Cross-check: Epoch AI's independent 1 GW model lands at $38B for terrestrial; Bernstein $35B, Nvidia $50-60B, Orennia $60B. The orbital figure has no external benchmark — it is built from SpaceX's own disclosed specs.")
row += 1
b_ratio = row
ws.cell(row=row, column=1, value="  Orbital as a multiple of terrestrial").font = lbl_font(italic=True, size=9)
ws.cell(row=row, column=2, value="").font = lbl_font()
c = ws.cell(row=row, column=3, value=f"=C{b_total}/B{b_total}"); c.font = formula_font(size=9, bold=True); c.number_format = '0.00"x"'
row += 2

# ---------------- ANNUALIZED TCO ----------------
section_row(ws, row, "ANNUALIZED COST OF OWNERSHIP PER GW (where free energy actually shows up)", ncols=3); row += 1
r_it_yrs = two_col_input(ws, row, "Compute hardware life (years)", 5, 5, NUM0,
    "Epoch AI assumes 5 years for IT equipment on the ground. Assumed equal in orbit — chip obsolescence, not wear, is the binding constraint."); row += 1
r_inf_yrs = two_col_input(ws, row, "Infrastructure life (years)", 14, 5, NUM0,
    "THE CRITICAL ASYMMETRY. A terrestrial building, substation and chiller plant serve ~14 years and multiple server refreshes. An orbital satellite cannot be serviced — refreshing the compute means relaunching the entire spacecraft, so ALL of it is written off on the compute cycle."); row += 1
r_energy = two_col_input(ws, row, "Annual energy cost ($mm per GW per year)", 600, 0, CUR0,
    "Epoch AI puts energy at ~$0.6B/yr for a 1 GW facility. Zero in orbit — this is the headline 'free energy' saving."); row += 1
r_otheropex = two_col_input(ws, row, "Other annual opex ($mm per GW per year)", 300, 150, CUR0,
    "Terrestrial: water, maintenance, staffing. Orbital: ground stations and constellation operations. Musk claims 'no ongoing operational or maintenance needs' in orbit; a residual is assumed here."); row += 1

row += 1
t_itann = row
ws.cell(row=row, column=1, value="Annualized compute hardware").font = lbl_font()
for j, col in enumerate(["B", "C"]):
    f = f"={col}{b_compute}/{col}{r_it_yrs}"
    c = ws.cell(row=row, column=2+j, value=f); c.font = formula_font(); c.number_format = CUR0; c.border = box
row += 1
t_infann = row
ws.cell(row=row, column=1, value="Annualized infrastructure (all non-compute capex)").font = lbl_font()
for j, col in enumerate(["B", "C"]):
    f = f"=({col}{b_total}-{col}{b_compute})/{col}{r_inf_yrs}"
    c = ws.cell(row=row, column=2+j, value=f); c.font = formula_font(); c.number_format = CUR0; c.border = box
nrow(ws, row, "Terrestrial spreads this over 14 years; orbital over 5. This single line drives most of the result.")
row += 1
t_en = row
ws.cell(row=row, column=1, value="Energy").font = lbl_font()
for j, col in enumerate(["B", "C"]):
    c = ws.cell(row=row, column=2+j, value=f"={col}{r_energy}"); c.font = link_font(); c.number_format = CUR0; c.border = box
row += 1
t_ot = row
ws.cell(row=row, column=1, value="Other opex").font = lbl_font()
for j, col in enumerate(["B", "C"]):
    c = ws.cell(row=row, column=2+j, value=f"={col}{r_otheropex}"); c.font = link_font(); c.number_format = CUR0; c.border = box
row += 1
t_tot = row
ws.cell(row=row, column=1, value="TOTAL ANNUALIZED COST PER GW").font = Font(name=FONT, size=11, bold=True)
for j, col in enumerate(["B", "C"]):
    c = ws.cell(row=row, column=2+j, value=f"=SUM({col}{t_itann}:{col}{t_ot})")
    c.font = Font(name=FONT, size=11, bold=True); c.number_format = CUR0; c.border = dbl_top
row += 1
t_ratio = row
ws.cell(row=row, column=1, value="  Orbital as a multiple of terrestrial").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=3, value=f"=C{t_tot}/B{t_tot}"); c.font = formula_font(size=9, bold=True); c.number_format = '0.00"x"'
row += 1
t_be = row
ws.cell(row=row, column=1, value="  Orbital satellite life needed to match terrestrial (years)").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=3,
    value=f"=(C{b_total}-C{b_compute})/(B{t_tot}-C{r_energy}-C{r_otheropex}-C{b_compute}/C{r_it_yrs})")
c.font = formula_font(size=9, bold=True); c.number_format = '0.0" yrs"'
nrow(ws, row, "Holding everything else fixed, this is the on-orbit service life at which orbital annualized cost equals terrestrial. Longer-lived satellites are the single biggest lever — bigger than launch cost.")
row += 2

# ---------------- SENSITIVITY ----------------
section_row(ws, row, "SENSITIVITY: ORBITAL CAPEX PER GW ($mm)", ncols=7); row += 1
sens_hdr = row
ws.cell(row=row, column=1, value="Power density (kW/ton)  \\  Launch cost ($/kg)").font = hdr_font(size=10, color="000000")
ws.cell(row=row, column=1).fill = grey_fill; ws.cell(row=row, column=1).border = box
launch_vals = [10, 100, 200, 500, 1000]
for j, lv in enumerate(launch_vals):
    c = ws.cell(row=row, column=2+j, value=lv)
    c.font = hdr_font(size=10, color="000000"); c.fill = lblue_fill
    c.number_format = '$#,##0'; c.alignment = Alignment(horizontal="center"); c.border = box
row += 1
dens_vals = [50, 70, 100, 150]
for dv in dens_vals:
    c0 = ws.cell(row=row, column=1, value=dv)
    c0.font = hdr_font(size=10, color="000000"); c0.fill = lblue_fill; c0.number_format = NUM0; c0.border = box
    for j in range(len(launch_vals)):
        lcol = get_column_letter(2+j)
        f = f"=(C{b_total}-C{b_launch})+(1000000/$A{row})*{lcol}${sens_hdr}/1000"
        c = ws.cell(row=row, column=2+j, value=f)
        c.font = formula_font(); c.number_format = CUR0; c.border = box
    row += 1
sens_note = row
ws.cell(row=row, column=1, value="Only the launch component varies in this grid; all other orbital costs are fixed by satellite count, which is set by kW per satellite rather than kW per ton. Terrestrial comparison point is the total above.").font = lbl_font(size=8, italic=True, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

# ---------------- READ-ACROSS ----------------
section_row(ws, row, "READ-ACROSS", ncols=3); row += 1
for txt in [
 "Launch cost is not the binding constraint. At the $200/kg Starship target, getting a full gigawatt of compute to orbit costs only a few billion dollars — a modest share of the total. Even at Musk's aspirational $10/kg, the orbital total barely moves, because the cost sits in the hardware, not the ride.",
 "Free energy and no water cooling are real but second-order. They remove the grid, backup generation and chiller plant from capex, and roughly $0.6B/GW/year of energy opex. Against that, orbital must add solar arrays, deployable radiators, a satellite bus and launch — which together cost more than what they replace.",
 "The decisive variable is service life. A terrestrial data centre amortizes its shell, power and cooling over ~14 years across several server refreshes. A satellite cannot be serviced, so refreshing compute means rebuilding and relaunching everything. Writing the whole spacecraft off on a 5-year compute cycle is what makes orbital more expensive per year, despite the free power.",
 "Implication for the forecast: the $55B/GW orbital placeholder used on the Assumptions tab is broadly consistent with this bottom-up build, so the model's orbital memo block is not obviously mis-sized. But orbital does not look cheaper than terrestrial on these inputs, which is the opposite of the case usually made for it.",
 "Caveats: no external benchmark exists for orbital capex per GW — every figure on that side is derived from SpaceX's own disclosed specs and reasonable unit-cost assumptions. Several independent analyses have questioned whether AI1's disclosed radiator area can reject the heat its stated compute payload generates at all; if it cannot, radiator cost and mass both rise materially and the comparison worsens further.",
]:
    c = ws.cell(row=row, column=1, value=txt)
    c.font = lbl_font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 42
    row += 1

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Terrestrial vs Orbital sheet done")

# =========================================================================
# 7c. SCENARIOS TO 2030
# =========================================================================
ws = wb.create_sheet("Scenarios to 2030")
sheet_header(ws, "Revenue Scenarios to 2030 — What Would Have To Be True",
    "$ in billions, calendar years. Each scenario takes a revenue path as the input and solves BACKWARDS for the compute, capex, funding and launch cadence required to deliver it.")
set_col_widths(ws, [46, 12, 12, 12, 12, 12, 3, 82])

YRS = ["2026E", "2027E", "2028E", "2029E", "2030E"]
YC = ["B", "C", "D", "E", "F"]
NOTE_COL = 8
BN = '$#,##0.0,,"B"'
BN0 = '$#,##0'

def snote(ws, row, text):
    c = ws.cell(row=row, column=NOTE_COL, value=text)
    c.font = lbl_font(size=8, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")

row = 4
col_headers(ws, row, ["$B (calendar year)"] + YRS)
c = ws.cell(row=row, column=NOTE_COL, value="Basis / source")
c.font = hdr_font(size=10, color="000000"); c.fill = grey_fill; c.border = box
row += 1

# ---------- SHARED UNIT ECONOMICS ----------
section_row(ws, row, "SHARED UNIT ECONOMICS (editable — applied identically to every scenario)", ncols=6); row += 1

def yr_input(ws, row, label, vals, fmt, note_txt):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=2+i, value=v)
        c.font = input_font(); c.number_format = fmt; c.fill = yellow_fill; c.border = box
    snote(ws, row, note_txt)
    return row

u_revgw = yr_input(ws, row, "Revenue per revenue-generating GW ($B/yr)",
    [11.6, 11.0, 9.6, 9.0, 8.5], '$#,##0.0',
    "Carried from the Segment Forecast tab's recalibrated $/GW rate, which sits near CoreWeave's implied ~$10.9B/GW/yr. Declines as installed capacity outpaces contracted utilisation."); row += 1
u_monet = yr_input(ws, row, "Monetizable share of compute (%)",
    [0.90, 0.90, 0.90, 0.90, 0.90], PCT1,
    "Management indicated roughly 10% of compute goes to internal Grok training and earns no external revenue."); row += 1
u_capexgw = yr_input(ws, row, "Capex per incremental GW ($B/GW)",
    [30, 25, 22, 20, 19], BN0,
    "Recalibrated against Bernstein ($35B), Epoch AI ($38B), Nvidia ($50-60B) and CoreWeave's implied ~$46B/GW. Declines only gently — external sources describe per-GW build costs as flat-to-rising."); row += 1
u_nonai = yr_input(ws, row, "Non-AI capex as % of non-AI revenue",
    [0.40, 0.38, 0.35, 0.33, 0.32], PCT1,
    "Space plus Connectivity capex intensity, benchmarked to Q2'26A actuals."); row += 1
u_ebitda = yr_input(ws, row, "EBITDA margin (%)",
    [0.14, 0.20, 0.25, 0.28, 0.30], PCT1,
    "Used only to size the external funding gap. Q2'26A Adjusted EBITDA margin was 45%, but that predates the depreciation load from the build-out."); row += 1

row += 1
section_row(ws, row, "PHYSICAL CONSTANTS", ncols=6); row += 1
def const_input(ws, row, label, val, fmt, note_txt):
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    c = ws.cell(row=row, column=2, value=val)
    c.font = input_font(); c.number_format = fmt; c.fill = yellow_fill; c.border = box
    snote(ws, row, note_txt)
    return row
k_prior_gw = const_input(ws, row, "Nameplate compute at end-2025 (GW)", 1.0, '0.0',
    "SpaceX reported 1.0 GW at Q1'26A and 1.4 GW at Q2'26A."); row += 1
k_uspower = const_input(ws, row, "US average electricity demand (GW)", 460, NUM0,
    "Approximate average (not peak) US electricity draw, for scale only."); row += 1
k_globaldc = const_input(ws, row, "Global data centre capacity, all workloads (GW)", 60, NUM0,
    "Rough industry estimate of total installed capacity across all workloads, not just AI."); row += 1
k_tons = const_input(ws, row, "Mass to orbit per GW (tons)", 14286, NUM0,
    "From the Terrestrial vs Orbital tab: 1 GW divided by AI1's disclosed 70 kW/ton."); row += 1
k_payload = const_input(ws, row, "Starship payload to LEO per launch (tons)", 200, NUM0,
    "Block 4 design payload. Block 3 is 100t; using the higher figure is the generous assumption."); row += 1

row += 1

SCEN_ROWS = {}
def scenario_block(ws, row, title, rev_vals, ai_share_vals, header_note):
    section_row(ws, row, title, ncols=6); row += 1
    r = {}
    r['rev'] = yr_input(ws, row, "Total revenue ($B)  << INPUT", rev_vals, '$#,##0', header_note); row += 1
    r['aishare'] = yr_input(ws, row, "AI share of revenue (%)  << INPUT", ai_share_vals, PCT1,
        "Morgan Stanley attributes roughly $190B of its $330B 2030 estimate to AI; these paths assume AI carries a rising majority."); row += 1

    def calc(label, formula_fn, fmt, note_txt, bold=False, indent=False):
        nonlocal row
        lab = ("  " if indent else "") + label
        c0 = ws.cell(row=row, column=1, value=lab)
        c0.font = lbl_font(bold=bold, italic=indent, size=9 if indent else 10)
        cur = row
        for i, col in enumerate(YC):
            c = ws.cell(row=row, column=2+i, value=formula_fn(i, col, cur))
            c.font = formula_font(bold=bold, size=9 if indent else 10); c.number_format = fmt
            if bold: c.border = top_border
        snote(ws, row, note_txt)
        rr = row; row += 1
        return rr

    r['airev'] = calc("AI revenue ($B)", lambda i,c,_r=None: f"={c}{r['rev']}*{c}{r['aishare']}", '$#,##0',
        "")
    r['monetgw'] = calc("Revenue-generating compute required (GW)",
        lambda i,c,_r=None: f"={c}{r['airev']}/{c}{u_revgw}", '#,##0.0',
        "AI revenue divided by revenue per GW — this is the compute that must actually be earning.")
    r['gw'] = calc("NAMEPLATE COMPUTE REQUIRED (GW)",
        lambda i,c,_r=None: f"={c}{r['monetgw']}/{c}{u_monet}", '#,##0.0',
        "Grossed up for the internal Grok allocation. This is the physical capacity that must be built, powered and cooled.", bold=True)
    r['gwadd'] = calc("Incremental GW added in year",
        lambda i,c,_r=None: (f"={c}{r['gw']}-$B${k_prior_gw}" if i==0 else f"={c}{r['gw']}-{YC[i-1]}{r['gw']}"),
        '#,##0.0', "", indent=True)
    r['aicapex'] = calc("AI capex ($B)",
        lambda i,c,_r=None: f"=MAX(0,{c}{r['gwadd']})*{c}{u_capexgw}", '$#,##0', "")
    r['nonaicapex'] = calc("Non-AI capex ($B)",
        lambda i,c,_r=None: f"=({c}{r['rev']}-{c}{r['airev']})*{c}{u_nonai}", '$#,##0', "")
    r['capex'] = calc("Total capex ($B)",
        lambda i,c,_r=None: f"={c}{r['aicapex']}+{c}{r['nonaicapex']}", '$#,##0', "", bold=True)
    r['cumcapex'] = calc("Cumulative capex from 2026 ($B)",
        lambda i,c,_r=None: (f"={c}{r['capex']}" if i==0 else f"={YC[i-1]}{_r}+{c}{r['capex']}"),
        '$#,##0', "", indent=True)
    r['ebitda'] = calc("EBITDA ($B)",
        lambda i,c,_r=None: f"={c}{r['rev']}*{c}{u_ebitda}", '$#,##0', "")
    r['fund'] = calc("CUMULATIVE EXTERNAL FUNDING NEEDED ($B)",
        lambda i,c,_r=None: (f"=MAX(0,{c}{r['capex']}-{c}{r['ebitda']})" if i==0
                     else f"=MAX(0,{YC[i-1]}{_r}+{c}{r['capex']}-{c}{r['ebitda']})"),
        '$#,##0', "Capex less EBITDA, cumulated. Ignores working capital and existing cash, so it understates rather than overstates the need.", bold=True)
    r['uspow'] = calc("Nameplate GW as % of US average power demand",
        lambda i,c,_r=None: f"={c}{r['gw']}/$B${k_uspower}", PCT1, "", indent=True)
    r['dcx'] = calc("Nameplate GW vs today's global data centre capacity",
        lambda i,c,_r=None: f"={c}{r['gw']}/$B${k_globaldc}", '0.00"x"', "", indent=True)
    r['launch'] = calc("Starship launches per DAY if built in orbit",
        lambda i,c,_r=None: f"=MAX(0,{c}{r['gwadd']})*$B${k_tons}/$B${k_payload}/365", '#,##0.0',
        "Feasibility check only — assumes the year's incremental capacity is delivered to orbit rather than on the ground. SpaceX has flown 78 launches so far in 2026, about 0.4 per day.", indent=True)
    row += 1
    return row, r

row, SCEN_ROWS['base'] = scenario_block(ws, row,
    "SCENARIO A — BASE (this model, extended)",
    [36, 98, 169, 236, 307], [0.55, 0.68, 0.72, 0.75, 0.76],
    "2026-2028 are this workbook's own quarterly forecast summed to calendar years. 2029-2030 extend it at decelerating growth (+40%, +30%).")

row, SCEN_ROWS['street'] = scenario_block(ws, row,
    "SCENARIO B — STREET (Goldman Sachs / Morgan Stanley)",
    [36, 75, 160, 207, 400], [0.55, 0.68, 0.72, 0.75, 0.76],
    "Morgan Stanley models $160B in 2028 and $330B in 2030; Goldman Sachs $470B in 2030; FactSet consensus is near $207B for 2029. 2030 shown at the $400B midpoint of the two banks.")

row, SCEN_ROWS['musk'] = scenario_block(ws, row,
    "SCENARIO C — MUSK ($1 TRILLION BY 2030)",
    [36, 130, 290, 600, 1000], [0.55, 0.72, 0.78, 0.80, 0.80],
    "Musk told analysts on the Q2'26 call that internal projections for $1 trillion of revenue — 'not ARR, but revenue' — moved from 2031 to 2030, with a non-zero chance of 2029. Intermediate years are a smooth path fitted to that endpoint.")

# ---------- SUMMARY ----------
section_row(ws, row, "2030 SIDE-BY-SIDE", ncols=6); row += 1
sum_hdr = row
for j, h in enumerate(["", "Base", "Street", "Musk", "", ""]):
    c = ws.cell(row=row, column=1+j, value=h)
    if j > 0 and h:
        c.font = hdr_font(size=10, color="000000"); c.fill = lblue_fill
        c.alignment = Alignment(horizontal="center"); c.border = box
row += 1
for label, key, fmt in [
    ("Revenue ($B)", 'rev', '$#,##0'),
    ("Nameplate compute required (GW)", 'gw', '#,##0.0'),
    ("Capex in 2030 alone ($B)", 'capex', '$#,##0'),
    ("Cumulative capex 2026-2030 ($B)", 'cumcapex', '$#,##0'),
    ("Cumulative external funding needed ($B)", 'fund', '$#,##0'),
    ("Share of US average power demand", 'uspow', PCT1),
    ("Multiple of global data centre capacity", 'dcx', '0.00"x"'),
    ("Starship launches per day if orbital", 'launch', '#,##0.0'),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font(bold=(key in ('rev','gw','fund')))
    for j, sc in enumerate(['base', 'street', 'musk']):
        c = ws.cell(row=row, column=2+j, value=f"=F{SCEN_ROWS[sc][key]}")
        c.font = formula_font(bold=(key in ('rev','gw','fund'))); c.number_format = fmt; c.border = box
    row += 1

row += 1
section_row(ws, row, "READ-ACROSS", ncols=6); row += 1
for txt in [
 "This model is not conservative relative to the people who underwrote the IPO. Its 2028 revenue of $169B sits above Morgan Stanley's $160B, and its exit run-rate implies a 2029 above the ~$207B FactSet consensus. It is conservative only against Musk's own aspiration, which both lead banks model at roughly a third.",
 "The $1 trillion case does not fail on financial assumptions — it fails on physics. Delivering it requires roughly 90 GW of nameplate compute, which is around a fifth of average total US electricity demand and more than the entire world's installed data centre capacity across all workloads today.",
 "Capital is the second binding constraint. The Musk path needs on the order of $1.5 trillion of cumulative capex between 2026 and 2030, against a company that raised $86 billion at IPO and whose shares fell 14% when quarterly capex came in at $18 billion.",
 "The orbital route does not relieve the constraint, it relocates it. Building the incremental capacity in space at AI1's disclosed 70 kW/ton implies a Starship cadence measured in dozens of launches per day, against roughly 0.4 per day today. And per the Terrestrial vs Orbital tab, orbital costs more per GW-year, not less.",
 "How to use this tab: the revenue rows are inputs and everything below them is solved. Rather than debating whether a revenue number is too low, change it and read off what it demands in gigawatts, dollars and launches — then judge whether that is deliverable. The unit economics are deliberately held identical across all three scenarios so the comparison isolates scale, not assumptions.",
]:
    c = ws.cell(row=row, column=1, value=txt)
    c.font = lbl_font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 40
    row += 1

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Scenarios to 2030 sheet done")

# =========================================================================
# 7d. MODEL vs BRIAN WANG (NEXTBIGFUTURE)
# =========================================================================
ws = wb.create_sheet("Model vs Brian Wang")
sheet_header(ws, "This Model vs Brian Wang / NextBigFuture",
    "$ in billions. Wang's published SpaceX calls compared to this model, plus the bridge showing exactly what would have to change to reach them. His Tesla calls are shown separately for context — they are not comparable to a SpaceX model line-by-line.")
set_col_widths(ws, [46, 15, 15, 15, 3, 78])

NC = 6
def wnote(ws, row, text):
    c = ws.cell(row=row, column=NC, value=text)
    c.font = lbl_font(size=8, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")

row = 4
col_headers(ws, row, ["Metric", "This model", "Brian Wang", "Difference"])
c = ws.cell(row=row, column=NC, value="Source / comment")
c.font = hdr_font(size=10, color="000000"); c.fill = grey_fill; c.border = box
row += 1

# ---------- SCORECARD ----------
section_row(ws, row, "SPACEX — HEAD TO HEAD (directly comparable)", ncols=4); row += 1

def cmp_row(ws, row, label, model_v, wang_v, fmt, note_txt, diff="ratio", bold=False):
    ws.cell(row=row, column=1, value=label).font = lbl_font(bold=bold)
    c1 = ws.cell(row=row, column=2, value=model_v)
    c1.font = (formula_font(bold=bold) if isinstance(model_v, str) and model_v.startswith("=") else input_font(bold=bold))
    c1.number_format = fmt; c1.border = box
    c2 = ws.cell(row=row, column=3, value=wang_v)
    c2.font = input_font(bold=bold); c2.number_format = fmt; c2.border = box
    c3 = ws.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},\"n/a\")")
    c3.font = formula_font(bold=bold); c3.number_format = '0.00"x"'; c3.border = box
    wnote(ws, row, note_txt)
    return row

r_arr26 = cmp_row(ws, row, "End-2026 revenue run-rate ($B)", 51.2, 82,
    '$#,##0.0',
    "Wang's published base case was $82B ARR for end-2026, with a bull case near $96B. SpaceX then guided to $100B+ ARR by December — so Wang was LOW, and this model is lower still. Note ARR (contracted run-rate) and GAAP revenue are not the same measure; this comparison is indicative.", bold=True); row += 1
r_run27 = cmp_row(ws, row, "End-2027 revenue run-rate ($B)", 139.6, 300,
    '$#,##0.0',
    "Wang: '$300 billion per year by end of 2027', and he notes SemiAnalysis reached the same figure independently. This model's Q4'27E revenue annualised is $139.6B — about 47% of his number.", bold=True); row += 1
r_2030 = cmp_row(ws, row, "2030 revenue ($B)", 307, 1000,
    '$#,##0',
    "Wang has SpaceX above $1 trillion in 2030, matching Musk's own revised target. Morgan Stanley is at $330B and Goldman Sachs at $470B. Model figure is the Base scenario from the Scenarios to 2030 tab.", bold=True); row += 1
r_gw27 = cmp_row(ws, row, "Nameplate compute at end-2027 (GW)", 10.0, 10.0,
    '#,##0.0',
    "This model reaches 10.0 GW at Q4'27E, and Wang and SemiAnalysis both work from a roughly 10 GW end-2027 base. The two sides AGREE on capacity — which is what makes the revenue gap below so informative."); row += 1
r_revgw = cmp_row(ws, row, "Implied revenue per nameplate GW ($B/yr)",
    "=B" + "%d" % (row-3) + "*0.72/B" + "%d" % (row-1),
    "=C" + "%d" % (row-3) + "*0.72/C" + "%d" % (row-1),
    '$#,##0.0',
    "THE CRUX. Both sides assume a similar gigawatt build-out, so the entire revenue gap resolves into monetisation per GW. Wang's $300B on ~10 GW implies roughly $21.6B of revenue per nameplate GW per year; this model assumes about $10B, anchored to CoreWeave's disclosed economics.", bold=True); row += 1

row += 1
# ---------- BRIDGE ----------
section_row(ws, row, "BRIDGE — WHAT WOULD HAVE TO CHANGE TO REACH $300B BY END-2027", ncols=4); row += 1
b_start = row
ws.cell(row=row, column=1, value="This model, Q4'27E annualised").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"=B{r_run27}"); c.font = link_font(); c.number_format = '$#,##0.0'; c.border = box
wnote(ws, row, "Starting point."); row += 1
b_gw = row
ws.cell(row=row, column=1, value="Required nameplate GW at THIS model's $/GW").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"=C{r_run27}*0.72/(B{r_revgw})"); c.font = formula_font(); c.number_format = '#,##0.0'; c.border = box
wnote(ws, row, "Holding monetisation at this model's rate, Wang's $300B would require this much capacity — versus the 10 GW he and SemiAnalysis actually assume. This route is not physically available on his own capacity numbers."); row += 1
b_rate = row
ws.cell(row=row, column=1, value="Required $/GW at Wang's 10 GW ($B/yr)").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"=C{r_run27}*0.72/C{r_gw27}"); c.font = formula_font(); c.number_format = '$#,##0.0'; c.border = box
wnote(ws, row, "The other route: keep capacity as modelled and lift the rent. This is roughly 2x CoreWeave's implied ~$10.9B/GW/yr and about 3.4x SpaceX's own Q2'26A realised rate of ~$6.3B/GW/yr."); row += 1
b_mult = row
ws.cell(row=row, column=1, value="  Multiple of CoreWeave's implied realised rate").font = lbl_font(italic=True, size=9)
c = ws.cell(row=row, column=2, value=f"=B{b_rate}/10.9"); c.font = formula_font(size=9, bold=True); c.number_format = '0.00"x"'; c.border = box
wnote(ws, row, "CoreWeave FY26 guidance: $18-19B exit ARR on ~1.7 GW of active power."); row += 1
b_capex = row
ws.cell(ledger := row, column=1, value="Capex implied by Wang's 10 GW at this model's $/GW ($B)").font = lbl_font()
c = ws.cell(row=row, column=2, value=f"=(C{r_gw27}-1.4)*25"); c.font = formula_font(); c.number_format = '$#,##0'; c.border = box
wnote(ws, row, "Incremental GW from Q2'26A's 1.4 GW, at $25B/GW. Broadly consistent with this model's own 2027 capex, so the two sides do not differ much on spending — only on what that spending earns."); row += 1

row += 1
# ---------- TESLA / MERGED ----------
section_row(ws, row, "WANG'S TESLA AND MERGED-ENTITY CALLS (context only — not comparable to this model)", ncols=4); row += 1
for label, val, note_txt in [
    ("Tesla revenue, 2030", "Above $1 trillion",
     "From Wang's June 2026 piece 'SpaceX and Tesla Revised Revenue and Earnings; Both Over $1T in 2030'. Driven by robotaxi, Optimus and energy storage — none of which this SpaceX model contains, so there is no line-by-line comparison to make."),
    ("Merged SpaceX + Tesla revenue, 2036", "$1.5-3+ quadrillion",
     "Predicated on Musk's '10x the world economy in 10 years' scenario, which Wang maps to ~27-29% global GDP CAGR. For scale, current world GDP is roughly $110 trillion, so this assumes the merged entity alone earns many times today's entire global output."),
    ("Merged entity market cap, 2036", "$300-1,000+ trillion",
     "Wang describes this as 'still only a few percent of the much larger global economy' under his 10x assumption. Today's entire global equity market is roughly $130 trillion."),
    ("Compute scaling from 2031", "100 GW/yr -> 1 TW/yr",
     "Wang's assumed build cadence beyond this model's horizon. For reference, the Scenarios to 2030 tab shows that even 105 GW of total installed capacity equates to roughly a fifth of US average electricity demand."),
    ("Assumed steady-state margins", "70%+",
     "Wang's rationale is that once capacity is built, incremental compute leasing carries very low marginal cost. This model reaches roughly 25-30% EBITDA margins, because depreciation on the build-out is charged explicitly — see the Financing & D&A tab."),
]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    c = ws.cell(row=row, column=2, value=val)
    c.font = input_font(); c.border = box
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c.alignment = Alignment(horizontal="center")
    wnote(ws, row, note_txt)
    row += 1

row += 1
# ---------- TRACK RECORD ----------
section_row(ws, row, "TRACK RECORD ON THE ONE FORECAST THAT HAS BEEN TESTED", ncols=4); row += 1
tr_hdr = row
for j, h in enumerate(["End-2026 ARR call", "Wang base", "Wang bull", "Company guide"]):
    c = ws.cell(row=row, column=1+j, value=h)
    c.font = hdr_font(size=10, color="000000"); c.fill = lblue_fill
    c.border = box; c.alignment = Alignment(horizontal="center" if j else "left")
row += 1
ws.cell(row=row, column=1, value="$B").font = lbl_font()
for j, v in enumerate([82, 96, 100]):
    c = ws.cell(row=row, column=2+j, value=v); c.font = input_font(); c.number_format = '$#,##0'; c.border = box
wnote(ws, row, "Wang's own retrospective: 'My projection of $82B ARR for end of 2026 was low.' The single testable data point so far went AGAINST him in the conservative direction — worth weighing in both directions when judging the $300B call.")
row += 2

# ---------- READ-ACROSS ----------
section_row(ws, row, "READ-ACROSS", ncols=4); row += 1
for txt in [
 "The disagreement is not about gigawatts — it is about rent. Wang, SemiAnalysis and this model all work from a roughly 10 GW end-2027 base, and the implied capex is similar. The entire gap between $140B and $300B resolves into revenue per gigawatt: about $10B/GW/yr here versus roughly $21.6B implied by Wang.",
 "His case for the higher rate is specific and not unreasonable. Wang argues SpaceX takes premium spot pricing now rather than locking into the multi-year fixed contracts CoreWeave and Nebius use, and can re-let capacity or self-consume it through Grok and Cursor. If AI compute stays supply-constrained, realised rates could exceed the contracted benchmarks this model is anchored to.",
 "The case against is that no operator has yet demonstrated it. This model's $10B/GW/yr already sits at CoreWeave's disclosed level and above SpaceX's own realised Q2'26A rate of roughly $6.3B/GW/yr. Wang's figure requires roughly doubling the best rate any public comparable currently achieves, and holding it while capacity grows sevenfold.",
 "Track record cuts both ways. Wang's one testable call — end-2026 ARR — came in low, not high, and SpaceX beat it. That argues against dismissing the $300B figure as pure enthusiasm. It does not, on its own, validate a monetisation rate twice the best observed benchmark.",
 "The Tesla and merged-entity numbers are a different kind of claim. They rest on Musk's 10x-world-economy premise rather than on company-specific unit economics, and the quadrillion-scale figures exceed current global GDP by more than an order of magnitude. They are not testable against a SpaceX P&L model and are reproduced here only so the full set of Wang's published calls is visible in one place.",
 "How to use this tab: if you find Wang's monetisation argument persuasive, the lever is the 'AI infrastructure revenue per GW' row on the Assumptions tab. Raising it toward $5,400/GW/quarter reproduces his $300B end-2027 figure without touching the capacity build — and will also shorten the implied capex payback, which is the sanity check worth watching.",
]:
    c = ws.cell(row=row, column=1, value=txt)
    c.font = lbl_font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 46
    row += 1

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Model vs Brian Wang sheet done")

# =========================================================================
# 7e. CASE COMPARISON & SENSITIVITY
# =========================================================================
ws = wb.create_sheet("Case Comparison")
sheet_header(ws, "Case Comparison & Sensitivity",
    "All three cases computed live from the case library, independent of which case is currently selected. Change a case-library input on the Assumptions tab and every column here updates.")
set_col_widths(ws, [46, 15, 15, 15, 3, 76])

CN = 6
def cnote(ws, row, text):
    c = ws.cell(row=row, column=CN, value=text)
    c.font = lbl_font(size=8, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True, vertical="top")

SFQ = lambda i: SF_FCOLS[i]     # Segment Forecast forecast column
AQ  = lambda i: ASSUMP_COLS[i]  # Assumptions forecast column
CPQ = lambda i: CP_COLS[3+i]    # Consolidated P&L forecast column
SFS = "'Segment Forecast'!"
CPS = "'Consolidated P&L'!"
AS_ = "Assumptions!"

def infra(case, i):
    return f"{SFS}{SFQ(i)}{SF_ROWS['rev_gw']}*{AS_}{AQ(i)}{LIB[(case,'blended')]}"
def totrev(case, i):
    return (f"({infra(case,i)}+{SFS}{SFQ(i)}{SF_ROWS['grok']}+{SFS}{SFQ(i)}{SF_ROWS['cursor']}"
            f"+{SFS}{SFQ(i)}{SF_ROWS['adv']}+{CPS}{CPQ(i)}{CP_ROWS['space_rev']}+{CPS}{CPQ(i)}{CP_ROWS['conn_rev']})")
def arr(case, i):
    return f"{totrev(case,i)}*{AS_}$B${ASSUMP_ROWS['dec_share']}*12"
def aicapex(case, i):
    return f"MAX(0,{SFS}{SFQ(i)}{SF_ROWS['gw_added']})*{AS_}{AQ(i)}{LIB[(case,'capex')]}"
def totcapex(case, i):
    return f"({aicapex(case,i)}+{CPS}{CPQ(i)}{CP_ROWS['space_capex']}+{CPS}{CPQ(i)}{CP_ROWS['conn_capex']})"
def payback(case, i):
    return f"{AS_}{AQ(i)}{LIB[(case,'capex')]}/({AS_}{AQ(i)}{LIB[(case,'marginal')]}*4)"

row = 4
hdr = ["Metric", "Case 1 Conservative", "Case 2 Mgmt-consistent", "Case 3 High-capex"]
col_headers(ws, row, hdr)
c = ws.cell(row=row, column=CN, value="Benchmark / comment")
c.font = hdr_font(size=10, color="000000"); c.fill = grey_fill; c.border = box
row += 1

r_active = row
ws.cell(row=row, column=1, value="Currently selected on Assumptions tab").font = lbl_font(bold=True)
for k in range(3):
    c = ws.cell(row=row, column=2+k, value=f'=IF({AS_}$B${ASSUMP_ROWS["case"]}={k+1},"<< ACTIVE","")')
    c.font = Font(name=FONT, size=10, bold=True, color="C00000"); c.border = box
    c.alignment = Alignment(horizontal="center")
row += 1

section_row(ws, row, "REVENUE RUN-RATE (management's exit-month x 12 basis)", ncols=4); row += 1
def metric(ws, row, label, fn, fmt, note_txt, bold=False):
    ws.cell(row=row, column=1, value=label).font = lbl_font(bold=bold)
    for k in range(3):
        c = ws.cell(row=row, column=2+k, value="=" + fn(k+1))
        c.font = formula_font(bold=bold); c.number_format = fmt; c.border = box
    cnote(ws, row, note_txt)
    return row

m_arr26 = metric(ws, row, "ARR at Dec 2026 ($mm)", lambda k: arr(k,1), CUR0,
    "Management guided to $100B+ ARR by December 2026, measured as expected December revenue annualised. Case 2 is calibrated to reproduce it.", bold=True); row += 1
m_arr27 = metric(ws, row, "ARR at Dec 2027 ($mm)", lambda k: arr(k,5), CUR0,
    "SemiAnalysis and Brian Wang both independently arrive at roughly $300B by end-2027.", bold=True); row += 1
m_arr28 = metric(ws, row, "ARR at Dec 2028 ($mm)", lambda k: arr(k,9), CUR0, ""); row += 1

row += 1
section_row(ws, row, "CAPITAL INTENSITY", ncols=4); row += 1
m_cx27 = metric(ws, row, "CY2027 total capex ($mm)",
    lambda k: "+".join(totcapex(k,i) for i in range(2,6)), CUR0,
    "SemiAnalysis-derived commentary implies $300-500B of capex in 2027 alone at roughly $50B per GW.", bold=True); row += 1
m_cx28 = metric(ws, row, "CY2028 total capex ($mm)",
    lambda k: "+".join(totcapex(k,i) for i in range(6,10)), CUR0, ""); row += 1
m_cxall = metric(ws, row, "Cumulative capex Q3'26E-Q4'28E ($mm)",
    lambda k: "+".join(totcapex(k,i) for i in range(10)), CUR0,
    "For scale: SpaceX raised about $86B net at IPO and held roughly $100B of cash and securities at Jun 30, 2026.", bold=True); row += 1

row += 1
section_row(ws, row, "RETURNS ON NEW CAPITAL", ncols=4); row += 1
m_pb26 = metric(ws, row, "Payback on new capacity, Q4'26E (yrs, mgmt basis)",
    lambda k: payback(k,1), '0.00"x"',
    "The CFO claimed 'less than a one-year payback' on new compute deployments. Case 2 lands at 1.04 years — close, though not strictly under one.", bold=True); row += 1
m_pb27 = metric(ws, row, "Payback on new capacity, Q4'27E (yrs, mgmt basis)",
    lambda k: payback(k,5), '0.00"x"', ""); row += 1
m_rate26 = metric(ws, row, "Blended revenue per GW, Q4'26E ($B/yr)",
    lambda k: f"{AS_}{AQ(1)}{LIB[(k,'blended')]}*4/1000", '$#,##0.0',
    "CoreWeave's FY26 guidance implies roughly $10.9B/GW/yr on contracted, non-spot terms."); row += 1
m_marg26 = metric(ws, row, "Marginal contract rate, Q4'26E ($B/yr)",
    lambda k: f"{AS_}{AQ(1)}{LIB[(k,'marginal')]}*4/1000", '$#,##0.0',
    "SemiAnalysis puts the Google agreement at roughly $48B per GW per year; Altimeter sees 2-3x neocloud pricing."); row += 1
m_cxgw26 = metric(ws, row, "Capex per GW, Q4'26E ($B)",
    lambda k: f"{AS_}{AQ(1)}{LIB[(k,'capex')]}/1000", '$#,##0.0',
    "Nvidia's Jensen Huang cites roughly $50B per GW today, rising toward $90B. Epoch AI models $38B; Bernstein $35B."); row += 1

row += 2
section_row(ws, row, "SENSITIVITY — PAYBACK ON NEW CAPACITY (years, management basis)", ncols=6); row += 1
sens_hdr = row
ws.cell(row=row, column=1, value="Capex per GW ($B)  \\  Marginal rate ($B/GW/yr)").font = hdr_font(size=10, color="000000")
ws.cell(row=row, column=1).fill = grey_fill; ws.cell(row=row, column=1).border = box
rate_vals = [12, 24, 36, 48, 60]
for j, rv in enumerate(rate_vals):
    c = ws.cell(row=row, column=2+j, value=rv)
    c.font = hdr_font(size=10, color="000000"); c.fill = lblue_fill
    c.number_format = '$#,##0'; c.alignment = Alignment(horizontal="center"); c.border = box
row += 1
capex_vals = [20, 30, 40, 50, 65, 90]
for cv in capex_vals:
    c0 = ws.cell(row=row, column=1, value=cv)
    c0.font = hdr_font(size=10, color="000000"); c0.fill = lblue_fill; c0.number_format = '$#,##0'; c0.border = box
    for j in range(len(rate_vals)):
        lcol = get_column_letter(2+j)
        c = ws.cell(row=row, column=2+j, value=f"=$A{row}/{lcol}${sens_hdr}")
        c.font = formula_font(); c.number_format = '0.00'; c.border = box
    row += 1
ws.cell(row=row, column=1, value="Values at or below 1.00 satisfy management's sub-1-year payback claim. The table shows it is achievable only where the marginal contract rate is at least roughly equal to capex per GW — which is precisely what the Google deal at ~$48B/GW against ~$50B/GW build cost implies, and precisely what Case 1's CoreWeave-anchored rates cannot deliver at any plausible build cost.").font = lbl_font(italic=True, size=8, color="808080")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 3

section_row(ws, row, "MODEL LIMITATIONS & HOW TO STRESS-TEST", ncols=4); row += 1
for txt in [
 "Every editable input in this workbook is a YELLOW cell. The ones that matter most, in order of impact: the case selector and case library on the Assumptions tab (AI capex per GW, blended revenue per GW, marginal rate); the nameplate GW schedule on the same tab; and the exit-month share that converts quarterly revenue into management's ARR measure.",
 "Grok and Cursor revenue are pure assumptions with no historical anchor. SpaceX does not disclose either separately — Grok sits inside the reported 'AI solutions and infrastructure' line and Cursor was not consolidated in Q2'26. They are modelled as standalone lines only because management stated the $100B ARR trajectory is 'including contribution from Cursor'. Treat their levels as illustrative.",
 "The cash flow statement is a proxy. It excludes working capital and deferred revenue movements, both of which have historically been meaningful sources of cash for SpaceX, and it assumes no equity issuance, no debt repayment and no asset sales. The funding gap it produces indicates scale, not a financing plan.",
 "Depreciation covers new forecast-period capex only. Depreciation on assets already in service at Q2'26 stays embedded in the cost ratios inherited from reported actuals, which avoids double-counting but means the D&A line is not a complete company depreciation figure.",
 "Orbital AI compute is deliberately excluded from every calculation and held as a memo on the Segment Forecast tab. It has no revenue history, no disclosed monetisation plan, and its capex per GW is derived from SpaceX's own specifications rather than any external benchmark. A one-click toggle to fold it into the forecast was considered and rejected: it would let speculative capacity flow into revenue, valuation and payback outputs that look authoritative but rest on no observable data.",
 "The three valuation blocks disagree by an order of magnitude and none should be relied on alone. Revenue-multiple gives $207-621B, PEG gives implausible trillions, and the DCF is dominated by a manually-entered terminal assumption. SPCX actually trades near $1.43T. The model is built to interrogate operating assumptions, not to produce a price target.",
]:
    c = ws.cell(row=row, column=1, value=txt)
    c.font = lbl_font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 48
    row += 1

ws.freeze_panes = "B5"
wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Case Comparison sheet done")

# =========================================================================
# 8. DASHBOARD
# =========================================================================
ws = wb.create_sheet("Dashboard")
sheet_header(ws, "Dashboard", "Revenue by segment, profitability trend, and AI compute build-out — actuals + forecast through Q4 2028")
set_col_widths(ws, [34] + [11]*len(ALLQ) + [3])

row = 4
col_headers(ws, row, ["$mm"] + ALLQ)
row += 1
dash_start = row
for label, cprow in [("Space revenue", cp_space_rev), ("Connectivity revenue", cp_conn_rev), ("AI revenue", cp_ai_rev)]:
    ws.cell(row=row, column=1, value=label).font = lbl_font()
    for i, col in enumerate(CP_COLS):
        cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{cprow}")
        cell.font = link_font(); cell.number_format = CUR0
    row += 1
dash_tot_rev = row
ws.cell(row=row, column=1, value="Total revenue").font = formula_font(bold=True)
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{cp_tot_rev}")
    cell.font = link_font(bold=True); cell.number_format = CUR0; cell.border = top_border
row += 1
dash_netloss = row
ws.cell(row=row, column=1, value="Net income (loss)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{cp_netloss}")
    cell.font = link_font(); cell.number_format = CUR0
row += 1
dash_gw = row
ws.cell(row=row, column=1, value="AI nameplate compute (GW) — terrestrial").font = lbl_font()
gw_hist = [None, None, 1.4]
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=gw_hist[i])
    else:
        fcol = SF_FCOLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"='Segment Forecast'!{fcol}{SF_ROWS['ai_gw']}")
        cell.font = link_font()
    cell.number_format = '0.0'
row += 1
dash_orb_gw = row
ws.cell(row=row, column=1, value="Orbital AI compute (GW) — speculative, SpaceX's own target").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i < 3:
        cell = ws.cell(row=row, column=2+i, value=0.0 if i == 2 else None)
    else:
        fcol = SF_FCOLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"='Segment Forecast'!{fcol}{SF_ROWS['orbital_gw']}")
        cell.font = link_font()
    cell.number_format = '0.0'
row += 1
dash_payback = row
ws.cell(row=row, column=1, value="Implied AI capex payback (years) vs. <1yr guidance").font = lbl_font()
for i, col in enumerate(CP_COLS):
    if i == 2:
        cell = ws.cell(row=row, column=2+i, value=f"='Segment Forecast'!B{SF_ROWS['ai_payback']}")
        cell.font = link_font()
    elif i >= 3:
        fcol = SF_FCOLS[i-3]
        cell = ws.cell(row=row, column=2+i, value=f"='Segment Forecast'!{fcol}{SF_ROWS['ai_payback']}")
        cell.font = link_font()
    else:
        cell = ws.cell(row=row, column=2+i, value=None)
    cell.number_format = '0.00" yrs"'
row += 1
dash_payback_target = row
ws.cell(row=row, column=1, value="  SpaceX guidance: <1-year payback").font = lbl_font(italic=True, size=9)
for i, col in enumerate(CP_COLS):
    if i >= 2:
        cell = ws.cell(row=row, column=2+i, value=1)
        cell.font = input_font(size=9)
    else:
        cell = ws.cell(row=row, column=2+i, value=None)
    cell.number_format = '0" yr"'
row += 1
dash_mktcap = row
ws.cell(row=row, column=1, value="Implied market cap ($B, revenue-multiple)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{CP_ROWS['mktcap']}")
    cell.font = link_font(); cell.number_format = '$#,##0.0"B"'
row += 1
dash_peg_mktcap = row
ws.cell(row=row, column=1, value="Implied market cap ($B, PEG-based)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{CP_ROWS['peg_mktcap']}")
    cell.font = link_font(); cell.number_format = '$#,##0.0"B";("n/m")'
row += 1
dash_capex = row
ws.cell(row=row, column=1, value="Total capex ($mm)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{CP_ROWS['tot_capex']}")
    cell.font = link_font(); cell.number_format = CUR0
row += 1
dash_ai_capex = row
ws.cell(row=row, column=1, value="  of which: AI capex ($mm)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"='Consolidated P&L'!{col}{CP_ROWS['ai_capex']}")
    cell.font = link_font(); cell.number_format = CUR0
row += 1
dash_other_capex = row
ws.cell(row=row, column=1, value="  of which: Space + Connectivity capex ($mm)").font = lbl_font()
for i, col in enumerate(CP_COLS):
    cell = ws.cell(row=row, column=2+i, value=f"={col}{dash_capex}-{col}{dash_ai_capex}")
    cell.font = formula_font(); cell.number_format = CUR0
row += 2

chart1 = BarChart()
chart1.type = "col"; chart1.grouping = "stacked"; chart1.overlap = 100
chart1.title = "Revenue by Segment ($mm) — Actuals & Forecast"
chart1.y_axis.title = "$mm"; chart1.style = 10
cats = Reference(ws, min_col=2, max_col=1+len(ALLQ), min_row=4, max_row=4)
data = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_start, max_row=dash_start+2)
chart1.add_data(data, titles_from_data=True, from_rows=True)
chart1.set_categories(cats)
chart1.width = 26; chart1.height = 10
ws.add_chart(chart1, f"A{row}")
row += 22

chart2 = LineChart()
chart2.title = "Total Revenue vs. Net Income (Loss) ($mm)"
chart2.y_axis.title = "$mm"; chart2.style = 12
data2 = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_tot_rev, max_row=dash_tot_rev)
data3 = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_netloss, max_row=dash_netloss)
chart2.add_data(data2, titles_from_data=True, from_rows=True)
chart2.add_data(data3, titles_from_data=True, from_rows=True)
chart2.set_categories(cats)
chart2.width = 26; chart2.height = 10
ws.add_chart(chart2, f"A{row}")
row += 22

chart3 = LineChart()
chart3.title = "AI Compute Build-out (GW) — Terrestrial (actual/forecast) vs. Orbital (speculative)"
chart3.y_axis.title = "GW"; chart3.style = 13
data4 = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_gw, max_row=dash_orb_gw)
chart3.add_data(data4, titles_from_data=True, from_rows=True)
chart3.set_categories(cats)
chart3.width = 26; chart3.height = 10
ws.add_chart(chart3, f"A{row}")
row += 22

chart3b = LineChart()
chart3b.title = "Implied AI Capex Payback (Years) vs. SpaceX's <1-Year Guidance"
chart3b.y_axis.title = "Years"; chart3b.style = 13
data4b = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_payback, max_row=dash_payback_target)
chart3b.add_data(data4b, titles_from_data=True, from_rows=True)
chart3b.set_categories(cats)
chart3b.width = 26; chart3b.height = 10
ws.add_chart(chart3b, f"A{row}")
row += 22

chart4 = LineChart()
chart4.title = "Implied Market Cap ($B) — Revenue-Multiple vs. PEG-Based Methodology"
chart4.y_axis.title = "$B"; chart4.style = 11
data5 = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_mktcap, max_row=dash_peg_mktcap)
chart4.add_data(data5, titles_from_data=True, from_rows=True)
chart4.set_categories(cats)
chart4.width = 26; chart4.height = 10
ws.add_chart(chart4, f"A{row}")
row += 22

chart5 = BarChart()
chart5.type = "col"; chart5.grouping = "stacked"; chart5.overlap = 100
chart5.title = "Total Capex ($mm) — AI Compute & Cooling vs. Space + Connectivity"
chart5.y_axis.title = "$mm"; chart5.style = 10
data6 = Reference(ws, min_col=1, max_col=1+len(ALLQ), min_row=dash_ai_capex, max_row=dash_other_capex)
chart5.add_data(data6, titles_from_data=True, from_rows=True)
chart5.set_categories(cats)
chart5.width = 26; chart5.height = 10
ws.add_chart(chart5, f"A{row}")

ws.freeze_panes = "B5"

order = ["Cover", "Assumptions", "KPIs", "Historical P&L", "Segment Forecast", "Consolidated P&L", FIN, "Balance Sheet & CF", "Terrestrial vs Orbital", "Scenarios to 2030", "Model vs Brian Wang", "Case Comparison", "Dashboard"]
wb._sheets = [wb[name] for name in order]
for name in order:
    wb[name].sheet_view.showGridLines = False

wb.save("/home/claude/spacex_model/spacex_model.xlsx")
print("Dashboard sheet done. All sheets:", wb.sheetnames)
